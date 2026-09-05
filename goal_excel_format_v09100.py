"""Target Excel display/rounding + previous-month actual prefill wiring.

- Remove trailing decimal points from numeric cells by using integer display format.
- Write Excel-only helper unit columns (G/I/M) as pre-calculated whole-won numbers,
  not formulas.
- v0.9.176 always applies the previous-month actual prefill patch to the freshly
  reloaded goal_excel_upload module before generating the workbook.
- v0.9.177 reloads that patch module first so an in-app update immediately uses
  the confirmed-quantity fix without requiring a full Python process restart.
"""
from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import importlib


def _num(value) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _whole(value) -> int:
    try:
        return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0


def _unit(total, qty) -> int:
    q = _num(qty)
    if abs(q) <= 1e-12:
        return 0
    return _whole(_num(total) / q)


def apply(upload_module):
    if upload_module is None:
        return upload_module

    # goal_data_status_v0985 reloads goal_excel_upload_v0984 on every render.
    # Apply the previous-month actual wrapper to that fresh module here, on the
    # exact export path, so the feature cannot exist as an unreferenced patch.
    prev = importlib.import_module("goal_prev_actual_template_v09174")
    prev = importlib.reload(prev)
    prev.apply(upload_module)

    if getattr(upload_module, "_rg_goal_excel_format_v09100_applied", False):
        return upload_module

    original = upload_module._template_bytes

    def template_bytes(core, db, month: str, base, old) -> bytes:
        from openpyxl import load_workbook

        data = original(core, db, month, base, old)
        bio = BytesIO(data)
        wb = load_workbook(bio)
        ws = wb["목표입력"] if "목표입력" in wb.sheetnames else wb.active

        # G/I/M are convenience values calculated once when the workbook is made.
        # They intentionally contain plain numbers rather than Excel formulas.
        numeric_cols = ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N")
        for row in range(2, ws.max_row + 1):
            qty = ws[f"E{row}"].value
            ws[f"G{row}"] = _unit(ws[f"F{row}"].value, qty)
            ws[f"I{row}"] = _unit(ws[f"H{row}"].value, qty)
            ws[f"M{row}"] = _unit(ws[f"L{row}"].value, qty)
            for col in numeric_cols:
                ws[f"{col}{row}"].number_format = "#,##0"

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    upload_module._template_bytes = template_bytes
    upload_module._rg_goal_excel_format_v09100_applied = True
    return upload_module
