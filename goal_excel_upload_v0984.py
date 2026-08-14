"""RG Manager Excel target template download/upload workflow.

The goal input tab is simplified to:
- download an Excel template populated with goal-managed active finished products
- keep the same product order as the goal/performance table
- fill target figures in Excel
- upload the workbook and save goals for the selected month by option ID

v0.9.94 excludes goal-management-excluded products from the template and upload save path.
v0.9.99 adds Excel-only unit helper columns for commission, RG logistics and COGS.
"""
from __future__ import annotations

from io import BytesIO
import importlib
import math
from typing import Any

import pandas as pd


_TEMPLATE_COLUMNS = [
    "아이템",
    "옵션ID",
    "매출",
    "단가",
    "수량",
    "수수료",
    "수수료단가",
    "입출고배송비",
    "입출고배송비단가",
    "반품처리비",
    "광고비",
    "상품원가",
    "상품원가단가",
    "매출이익",
]
_HELPER_COLUMNS = {
    "수수료단가",
    "입출고배송비단가",
    "상품원가단가",
}
_INPUT_COLUMNS = [
    "매출",
    "수량",
    "수수료",
    "입출고배송비",
    "반품처리비",
    "광고비",
    "상품원가",
    "매출이익",
]
_REQUIRED_UPLOAD_COLUMNS = ["아이템", "옵션ID"] + _INPUT_COLUMNS


def _clean_number(value: Any):
    try:
        if value is None or pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, str):
        s = value.replace(",", "").replace("원", "").replace("개", "").strip()
        if not s:
            return None
        value = s
    try:
        x = float(value)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def _goal_template_dataframe(core, db, month: str, base, old) -> pd.DataFrame:
    scope = importlib.import_module("goal_scope_v0994")
    products = scope.managed_products(core, db, base)
    goals = old._detail_goals(core, db, month, base)
    goal_map = {int(r["product_id"]): r for r in goals.to_dict("records")}

    product_records = list(products.itertuples(index=False)) if products is not None else []

    def _sort_key(p):
        pid = int(p.id)
        oid = base._oid(getattr(p, "option_id", "")) or base._oid(getattr(p, "item_code", ""))
        target_qty = old._num(goal_map.get(pid, {}).get("target_qty"))
        return (-target_qty, str(getattr(p, "name", "") or ""), str(oid or ""))

    product_records.sort(key=_sort_key)

    rows = []
    for p in product_records:
        pid = int(p.id)
        oid = base._oid(getattr(p, "option_id", "")) or base._oid(getattr(p, "item_code", ""))
        g = goal_map.get(pid)
        if g:
            revenue = old._num(g.get("target_revenue"))
            qty = old._num(g.get("target_qty"))
            commission = old._num(g.get("target_commission"))
            rg_cost = old._num(g.get("target_rg_cost"))
            cogs = old._num(g.get("target_cogs"))
            unit = revenue / qty if abs(qty) > 1e-12 else 0.0
            row = {
                "아이템": str(p.name or ""),
                "옵션ID": oid,
                "매출": revenue,
                "단가": unit,
                "수량": qty,
                "수수료": commission,
                "수수료단가": commission / qty if abs(qty) > 1e-12 else 0.0,
                "입출고배송비": rg_cost,
                "입출고배송비단가": rg_cost / qty if abs(qty) > 1e-12 else 0.0,
                "반품처리비": old._num(g.get("target_return_cost")),
                "광고비": old._num(g.get("target_ad_spend")),
                "상품원가": cogs,
                "상품원가단가": cogs / qty if abs(qty) > 1e-12 else 0.0,
                "매출이익": old._num(g.get("target_profit")),
            }
        else:
            row = {c: None for c in _TEMPLATE_COLUMNS}
            row["아이템"] = str(p.name or "")
            row["옵션ID"] = oid
        rows.append(row)
    return pd.DataFrame(rows, columns=_TEMPLATE_COLUMNS)


def _template_bytes(core, db, month: str, base, old) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    df = _goal_template_dataframe(core, db, month, base, old)
    wb = Workbook()
    ws = wb.active
    ws.title = "목표입력"

    header_fill = PatternFill("solid", fgColor="DCE8F6")
    input_fill = PatternFill("solid", fgColor="FFF7DF")
    helper_fill = PatternFill("solid", fgColor="EAF4EA")
    id_fill = PatternFill("solid", fgColor="F3F4F6")
    thin = Side(style="thin", color="CBD5E1")

    for col_idx, header in enumerate(_TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, rec in enumerate(df.to_dict("records"), 2):
        helper_formulas = {
            "수수료단가": f"=IFERROR(F{row_idx}/E{row_idx},0)",
            "입출고배송비단가": f"=IFERROR(H{row_idx}/E{row_idx},0)",
            "상품원가단가": f"=IFERROR(L{row_idx}/E{row_idx},0)",
        }
        for col_idx, header in enumerate(_TEMPLATE_COLUMNS, 1):
            value = helper_formulas.get(header, rec.get(header))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(header == "아이템"))
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if header in _HELPER_COLUMNS:
                cell.fill = helper_fill
            elif header in _INPUT_COLUMNS or header == "단가":
                cell.fill = input_fill
            else:
                cell.fill = id_fill
            if header in _INPUT_COLUMNS or header == "단가" or header in _HELPER_COLUMNS:
                cell.number_format = '#,##0.##'

    widths = {
        "A": 44,
        "B": 16,
        "C": 14,
        "D": 12,
        "E": 10,
        "F": 13,
        "G": 13,
        "H": 16,
        "I": 18,
        "J": 14,
        "K": 12,
        "L": 13,
        "M": 15,
        "N": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(1, ws.max_row)}"
    ws.sheet_view.showGridLines = False

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _row_has_input(row) -> bool:
    return any(_clean_number(row.get(col)) is not None for col in _INPUT_COLUMNS)


def _upload_rows(uploaded, base) -> pd.DataFrame:
    uploaded.seek(0)
    df = pd.read_excel(uploaded, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in _REQUIRED_UPLOAD_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError("필수 열이 없습니다: " + ", ".join(missing))
    # Excel-only helper columns are optional on upload so older target templates
    # remain compatible. They are display/calculation aids and are never saved.
    for col in _TEMPLATE_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[_TEMPLATE_COLUMNS].copy()
    df["옵션ID"] = df["옵션ID"].map(base._oid)
    df = df[df["옵션ID"].astype(str).str.strip() != ""]
    return df


def _save_uploaded_goals(core, db, month: str, df: pd.DataFrame, base, old):
    scope = importlib.import_module("goal_scope_v0994")
    all_products = base._products(core, db, active_only=True)
    excluded_pids = scope.excluded_ids(core, db)

    by_oid = {}
    excluded_by_oid = {}
    for p in all_products.itertuples(index=False):
        oid = base._oid(getattr(p, "option_id", "")) or base._oid(getattr(p, "item_code", ""))
        if not oid:
            continue
        pid = int(p.id)
        if pid in excluded_pids:
            excluded_by_oid[oid] = pid
        else:
            by_oid[oid] = pid

    seen = set()
    saved = 0
    skipped_blank = 0
    unknown = []
    excluded = []
    duplicates = []

    for _, r in df.iterrows():
        oid = base._oid(r.get("옵션ID"))
        if oid in seen:
            duplicates.append(oid)
            continue
        seen.add(oid)

        if oid in excluded_by_oid:
            excluded.append(oid)
            continue

        pid = by_oid.get(oid)
        if pid is None:
            unknown.append(oid)
            continue
        if not _row_has_input(r):
            skipped_blank += 1
            continue

        revenue = _clean_number(r.get("매출"))
        qty = _clean_number(r.get("수량"))
        commission = _clean_number(r.get("수수료"))
        rg = _clean_number(r.get("입출고배송비"))
        returns = _clean_number(r.get("반품처리비"))
        ad = _clean_number(r.get("광고비"))
        cogs = _clean_number(r.get("상품원가"))
        profit = _clean_number(r.get("매출이익"))

        revenue = revenue or 0.0
        qty = qty or 0.0
        commission = commission or 0.0
        rg = rg or 0.0
        returns = returns or 0.0
        ad = ad or 0.0
        cogs = cogs or 0.0
        if profit is None:
            profit = revenue - commission - rg - returns - ad - cogs

        payload = {
            "목표매출": revenue,
            "목표수량": qty,
            "목표수수료": commission,
            "목표입출고배송비": rg,
            "목표반품처리비": returns,
            "목표광고비": ad,
            "목표상품원가": cogs,
            "목표매출이익": profit,
            "메모": "",
        }
        old._save_detail_goal(core, db, month, pid, payload, base)
        saved += 1

    return {
        "saved": saved,
        "skipped_blank": skipped_blank,
        "unknown": sorted(set(unknown)),
        "excluded": sorted(set(excluded)),
        "duplicates": sorted(set(duplicates)),
    }


def _render_excel_goal_input(st, core, db, month: str, base, old):
    st.markdown("### 목표 엑셀 입력")
    st.caption(
        "양식을 내려받아 목표 숫자를 입력한 뒤 그대로 업로드하세요. "
        "목표·실적표와 같은 순서이며 옵션ID 기준으로 자동 저장됩니다. "
        "수수료단가·입출고배송비단가·상품원가단가는 엑셀에서 자동 계산되는 참고용 열입니다."
    )

    try:
        data = _template_bytes(core, db, month, base, old)
        st.download_button(
            "목표 엑셀 양식 다운로드",
            data=data,
            file_name=f"목표입력_{month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"goal984_download_{month}",
        )
    except Exception as exc:
        st.error(f"엑셀 양식을 만들지 못했습니다: {exc}")
        return

    uploaded = st.file_uploader(
        "작성한 목표 엑셀 업로드",
        type=["xlsx"],
        key=f"goal984_upload_{month}",
    )
    if uploaded is None:
        return

    try:
        df = _upload_rows(uploaded, base)
    except Exception as exc:
        st.error(f"엑셀을 읽지 못했습니다: {exc}")
        return

    st.success(f"엑셀 {len(df):,}행을 확인했습니다.")
    if st.button(
        f"{base._month_label(month)} 목표 저장",
        type="primary",
        use_container_width=True,
        key=f"goal984_save_{month}",
    ):
        try:
            result = _save_uploaded_goals(core, db, month, df, base, old)
        except Exception as exc:
            st.error(f"목표 저장 중 오류가 발생했습니다: {exc}")
            return

        st.success(f"목표 {result['saved']:,}개 상품을 저장했습니다.")
        if result["excluded"]:
            st.warning(
                "목표관리 제외 상품은 저장하지 않았습니다: "
                + ", ".join(result["excluded"][:20])
            )
        if result["unknown"]:
            st.warning("ERP에 없는 옵션ID: " + ", ".join(result["unknown"][:20]))
        if result["duplicates"]:
            st.warning("중복 옵션ID는 첫 행만 반영했습니다: " + ", ".join(result["duplicates"][:20]))
        st.rerun()


def render_page(st, pd_obj, core, db_path=None):
    base = importlib.import_module("goal_management_v0979")
    old = importlib.import_module("goal_excel_view_v0981")
    styled = importlib.import_module("goal_excel_view_v0983")
    db = db_path or core.DEFAULT_DB
    old._ensure_detail_schema(core, db, base)

    st.markdown(base._SELECT_CSS, unsafe_allow_html=True)
    st.markdown("## 🎯 목표·실적관리")
    st.caption("목표와 잠정실적·확정실적을 엑셀처럼 한 표에서 비교합니다.")

    months = base._month_options()
    month = st.selectbox(
        "목표·검증 월",
        months,
        index=0,
        format_func=base._month_label,
        key="goal_management_month_v0984",
    )

    tabs = st.tabs(["목표·실적표", "목표 입력", "월말검증", "목표이력"])
    with tabs[0]:
        styled._render_excel_comparison(st, core, db, month, base, old)
    with tabs[1]:
        _render_excel_goal_input(st, core, db, month, base, old)
    with tabs[2]:
        goals = base._goals(core, db, month)
        actuals, source_label = base._actuals(core, db, month)
        progress, _meta = base._build_progress(goals, actuals, month, core, db)
        base._render_review(st, core, db, month, progress, source_label)
    with tabs[3]:
        base._render_history(st, core, db)
