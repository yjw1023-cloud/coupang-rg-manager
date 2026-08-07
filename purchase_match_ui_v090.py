"""RG Manager v0.9.0 purchase matching review table.

Purpose
- Keep the existing purchase matching/import engine untouched.
- Show every Excel source item and its selected self-warehouse product in one table.
- Let the user correct a wrong match directly in that table.
- Feed the corrected product back into the original hidden selectbox so the existing
  confirmation/import path uses exactly the user's choice.

The original per-item expanders still execute for compatibility, but are hidden only
on the purchase page once the review table is ready.
"""
from __future__ import annotations

import hashlib
import io
import re
from typing import Any

from openpyxl import load_workbook

_CACHE_KEY = "_rg_purchase_match_cache_v090"
_OVERRIDE_KEY = "_rg_purchase_match_overrides_v090"
_READY_KEY = "_rg_purchase_match_ready_v090"
_APPLIED = False


def _primitive(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _extract_product_id(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, dict):
        for key in ("product_id", "id"):
            v = value.get(key)
            if isinstance(v, int):
                return int(v)
            if isinstance(v, float) and v.is_integer():
                return int(v)
            if isinstance(v, str) and v.isdigit():
                return int(v)
    for attr in ("product_id", "id"):
        try:
            v = getattr(value, attr)
        except Exception:
            v = None
        if isinstance(v, int):
            return int(v)
        if isinstance(v, float) and v.is_integer():
            return int(v)
        if isinstance(v, str) and v.isdigit():
            return int(v)
    if isinstance(value, (tuple, list)) and value:
        return _extract_product_id(value[0])
    return None


def _num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _fmt_qty(value: Any) -> str:
    n = _num(value)
    if abs(n - round(n)) < 1e-9:
        return f"{int(round(n)):,}"
    return f"{n:,.2f}".rstrip("0").rstrip(".")


def _fmt_money(value: Any) -> str:
    return f"{int(round(_num(value))):,}원"


def _file_bytes(uploaded: Any) -> bytes:
    if uploaded is None:
        return b""
    if hasattr(uploaded, "getvalue"):
        try:
            return uploaded.getvalue()
        except Exception:
            pass
    if isinstance(uploaded, (bytes, bytearray)):
        return bytes(uploaded)
    try:
        pos = uploaded.tell()
    except Exception:
        pos = None
    try:
        data = uploaded.read()
        if isinstance(data, str):
            data = data.encode("utf-8")
        return bytes(data or b"")
    finally:
        if pos is not None:
            try:
                uploaded.seek(pos)
            except Exception:
                pass


def _parse_purchase_excel(data: bytes) -> list[dict[str, Any]]:
    """Read the user-confirmed import columns: F/G source, W cost, AB quantity."""
    if not data:
        return []
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            for r in range(9, ws.max_row + 1):
                source_name = str(ws.cell(r, 6).value or "").strip()
                source_detail = str(ws.cell(r, 7).value or "").strip()
                unit_cost = ws.cell(r, 23).value
                qty = ws.cell(r, 28).value
                if not source_name:
                    continue
                if unit_cost in (None, "") or qty in (None, ""):
                    continue
                q = _num(qty)
                cost = _num(unit_cost)
                if q == 0:
                    continue
                rows.append({
                    "index": len(rows) + 1,
                    "sheet": ws.title,
                    "source_row": r,
                    "source_name": source_name,
                    "source_detail": source_detail,
                    "marking": str(ws.cell(r, 1).value or "").strip(),
                    "qty": q,
                    "unit_cost": cost,
                    "amount": q * cost,
                })
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return rows


def _self_warehouse_products(core_module, db_path) -> list[dict[str, Any]]:
    core_module.init_db(db_path)
    with core_module._conn(db_path) as c:
        rows = c.execute(
            """
            SELECT p.id, p.item_code, p.name, p.option_name, p.item_type,
                   COALESCE((
                       SELECT SUM(t.qty_delta)
                       FROM inventory_txns t
                       JOIN warehouses w ON w.id=t.warehouse_id
                       WHERE t.product_id=p.id AND w.name='자체창고'
                   ),0) AS own_stock,
                   EXISTS(
                       SELECT 1 FROM inventory_txns t
                       JOIN warehouses w ON w.id=t.warehouse_id
                       WHERE t.product_id=p.id AND w.name='자체창고'
                   ) AS own_history
            FROM products p
            WHERE p.active=1
              AND (
                   p.item_type='raw'
                   OR EXISTS(
                       SELECT 1 FROM inventory_txns t
                       JOIN warehouses w ON w.id=t.warehouse_id
                       WHERE t.product_id=p.id AND w.name='자체창고'
                   )
              )
            ORDER BY p.name, p.item_code
            """
        ).fetchall()
    out = []
    for r in rows:
        name = str(r["name"] or "").strip()
        opt = str(r["option_name"] or "").strip()
        code = str(r["item_code"] or "").strip()
        title = f"{name} [{opt}]" if opt and opt != name else name
        label = f"{title} | {code} | 자체재고 {_fmt_qty(r['own_stock'])}"
        out.append({
            "product_id": int(r["id"]),
            "item_code": code,
            "name": name,
            "option_name": opt,
            "own_stock": float(r["own_stock"] or 0),
            "label": label,
        })
    return out


def _parse_expander_label(label: str) -> tuple[int | None, str, str]:
    text = str(label or "").strip()
    m = re.match(r"^(\d+)\.\s*(.*?)\s*\|\s*([^|]+?)\s*$", text)
    if not m:
        return None, text, ""
    return int(m.group(1)), m.group(2).strip(), m.group(3).strip()


def _match_rate(text: str) -> str:
    m = re.search(r"(?:일치|유사도)\s*(\d+(?:\.\d+)?)\s*%", str(text or ""))
    if not m:
        return ""
    try:
        v = float(m.group(1))
        return f"{int(v)}%" if v.is_integer() else f"{v:g}%"
    except Exception:
        return f"{m.group(1)}%"


def _pid_from_display(display: str, products: list[dict[str, Any]]) -> int | None:
    text = str(display or "")
    for p in products:
        code = p["item_code"]
        if code and re.search(rf"(?<![A-Za-z0-9]){re.escape(code)}(?![A-Za-z0-9])", text):
            return int(p["product_id"])
    return None


def _safe_format(format_func, value: Any) -> str:
    try:
        return str(format_func(value)) if callable(format_func) else str(value)
    except Exception:
        return str(value)


def _set_original_widget_choice(st_obj, meta_row: dict[str, Any], product_id: int) -> bool:
    key = meta_row.get("widget_key")
    if not key:
        return False
    pid_to_value = meta_row.get("pid_to_value") or {}
    value = pid_to_value.get(str(int(product_id)), int(product_id))
    try:
        st_obj.session_state[key] = value
        return True
    except Exception:
        return False


def _render_review_table(st_obj, pd_obj, core_module, db_path, file_fp: str,
                         excel_rows: list[dict[str, Any]], meta: dict[str, Any]) -> None:
    products = _self_warehouse_products(core_module, db_path)
    if not products:
        st_obj.error("자체창고 매칭 후보 상품이 없습니다. 품목관리에서 자체창고 상품을 먼저 확인해 주세요.")
        return

    by_pid = {int(p["product_id"]): p for p in products}
    label_to_pid = {p["label"]: int(p["product_id"]) for p in products}
    product_labels = [p["label"] for p in products]
    placeholder = "— 매칭상품 선택 —"
    options = [placeholder] + product_labels

    overrides_all = st_obj.session_state.setdefault(_OVERRIDE_KEY, {})
    overrides = overrides_all.setdefault(file_fp, {})
    meta_rows = meta.get("rows") or {}

    table_rows = []
    current_pids: dict[int, int | None] = {}
    for row in excel_rows:
        idx = int(row["index"])
        mr = meta_rows.get(str(idx), {})
        current_pid = mr.get("current_pid")
        if current_pid is not None:
            try:
                current_pid = int(current_pid)
            except Exception:
                current_pid = None
        if current_pid is None:
            current_pid = _pid_from_display(mr.get("current_display", ""), products)
        selected_pid = overrides.get(str(idx), current_pid)
        try:
            selected_pid = int(selected_pid) if selected_pid is not None else None
        except Exception:
            selected_pid = current_pid
        current_pids[idx] = current_pid

        selected_label = by_pid.get(selected_pid, {}).get("label", placeholder)
        status_base = str(mr.get("status") or "")
        rate = str(mr.get("match_rate") or "")
        if str(idx) in overrides and selected_pid != current_pid:
            status = "수동 변경"
        elif "확인" in status_base:
            status = "확인 필요"
        elif "자동" in status_base:
            status = f"자동 {rate}".strip()
        elif selected_pid:
            status = rate or "매칭됨"
        else:
            status = "미매칭"

        source_display = row["source_name"]
        if row.get("source_detail"):
            source_display += f" · {row['source_detail']}"
        table_rows.append({
            "No.": idx,
            "매입상품": source_display,
            "수량": _fmt_qty(row["qty"]),
            "매입원가": _fmt_money(row["unit_cost"]),
            "매칭상품": selected_label,
            "상태": status,
        })

    auto_count = sum(1 for r in table_rows if str(r["상태"]).startswith("자동"))
    manual_count = sum(1 for r in table_rows if r["상태"] == "수동 변경")
    need_count = sum(1 for r in table_rows if r["상태"] in ("확인 필요", "미매칭"))
    c1, c2, c3 = st_obj.columns(3)
    c1.metric("자동매칭", f"{auto_count}개")
    c2.metric("수동수정", f"{manual_count}개")
    c3.metric("확인필요", f"{need_count}개")
    st_obj.caption("아래 표에서 매입상품과 매칭상품을 한 번에 확인하세요. 잘못 매칭된 행은 '매칭상품' 셀을 클릭해 자체창고 상품으로 바로 변경할 수 있습니다.")

    df = pd_obj.DataFrame(table_rows)
    editor_key = f"_rg_purchase_match_editor_v090_{file_fp[:16]}"
    column_config = None
    try:
        column_config = {
            "No.": st_obj.column_config.NumberColumn("No.", width="small"),
            "매입상품": st_obj.column_config.TextColumn("매입상품", width="large"),
            "수량": st_obj.column_config.TextColumn("수량", width="small"),
            "매입원가": st_obj.column_config.TextColumn("매입원가", width="small"),
            "매칭상품": st_obj.column_config.SelectboxColumn(
                "매칭상품",
                options=options,
                required=True,
                width="large",
                help="잘못 매칭되었으면 이 셀을 클릭해 올바른 자체창고 상품을 선택하세요.",
            ),
            "상태": st_obj.column_config.TextColumn("상태", width="small"),
        }
    except Exception:
        column_config = None

    editor_kwargs = dict(
        key=editor_key,
        hide_index=True,
        use_container_width=True,
        disabled=["No.", "매입상품", "수량", "매입원가", "상태"],
        num_rows="fixed",
    )
    if column_config is not None:
        editor_kwargs["column_config"] = column_config
    edited = st_obj.data_editor(df, **editor_kwargs)

    unresolved_widget = False
    for _, er in edited.iterrows():
        idx = int(er["No."])
        label = str(er.get("매칭상품") or placeholder)
        pid = label_to_pid.get(label)
        if pid is None:
            continue
        current_pid = current_pids.get(idx)
        if pid != current_pid:
            overrides[str(idx)] = pid
        else:
            overrides.pop(str(idx), None)

        mr = meta_rows.get(str(idx), {})
        target_pid = overrides.get(str(idx), current_pid)
        if target_pid is not None and not _set_original_widget_choice(st_obj, mr, int(target_pid)):
            unresolved_widget = True

    overrides_all[file_fp] = overrides
    st_obj.session_state[_OVERRIDE_KEY] = overrides_all
    if unresolved_widget:
        st_obj.warning("일부 오래된 매칭 위젯은 직접 연결키가 없어 수정값 반영을 확인할 수 없습니다. 프로그램을 최신 버전으로 다시 업데이트해 주세요.")
    elif overrides:
        st_obj.info("수동으로 바꾼 매칭은 아래 매입 확정 처리에 그대로 사용됩니다.")


def apply(purchase_module, core_module):
    """Patch purchase page UI while preserving matching/import business logic."""
    global _APPLIED
    if purchase_module is None:
        return purchase_module
    if getattr(purchase_module, "_rg_purchase_match_ui_v090_applied", False):
        return purchase_module

    original_render = getattr(purchase_module, "render_purchase_page", None)
    if not callable(original_render):
        purchase_module._rg_purchase_match_ui_v090_applied = True
        return purchase_module

    def render_purchase_page(*args, **kwargs):
        st_obj = kwargs.get("st")
        pd_obj = kwargs.get("pd")
        if st_obj is None:
            for obj in args:
                if hasattr(obj, "file_uploader") and hasattr(obj, "selectbox"):
                    st_obj = obj
                    break
        if pd_obj is None:
            try:
                import pandas as pd_obj
            except Exception:
                pd_obj = None
        if st_obj is None or pd_obj is None:
            return original_render(*args, **kwargs)

        db_path = core_module.DEFAULT_DB
        core_module.init_db(db_path)
        context: dict[str, Any] = {
            "current_index": None,
            "captured": {},
            "file_fp": None,
            "excel_rows": [],
            "file_name": "",
            "summary_rendered": False,
        }

        original_file_uploader = st_obj.file_uploader
        original_expander = st_obj.expander
        original_selectbox = st_obj.selectbox
        original_section = kwargs.get("section")

        class _TrackedExpander:
            def __init__(self, inner, idx, title, status):
                self._inner = inner
                self._idx = idx
                self._title = title
                self._status = status

            def __enter__(self):
                entered = self._inner.__enter__()
                context["current_index"] = self._idx
                if self._idx is not None:
                    row = context["captured"].setdefault(str(self._idx), {})
                    row["expander_title"] = self._title
                    row["status"] = self._status
                return entered

            def __exit__(self, exc_type, exc, tb):
                try:
                    return self._inner.__exit__(exc_type, exc, tb)
                finally:
                    context["current_index"] = None

            def __getattr__(self, name):
                return getattr(self._inner, name)

        def file_uploader_wrapper(*f_args, **f_kwargs):
            uploaded = original_file_uploader(*f_args, **f_kwargs)
            if uploaded is not None:
                data = _file_bytes(uploaded)
                if data:
                    fp = hashlib.sha1(data).hexdigest()
                    context["file_fp"] = fp
                    context["file_name"] = str(getattr(uploaded, "name", "") or "")
                    try:
                        context["excel_rows"] = _parse_purchase_excel(data)
                    except Exception:
                        context["excel_rows"] = []
            return uploaded

        def expander_wrapper(*e_args, **e_kwargs):
            label = str(e_args[0] if e_args else e_kwargs.get("label", ""))
            idx, title, status = _parse_expander_label(label)
            inner = original_expander(*e_args, **e_kwargs)
            return _TrackedExpander(inner, idx, title, status)

        def selectbox_wrapper(*s_args, **s_kwargs):
            result = original_selectbox(*s_args, **s_kwargs)
            idx = context.get("current_index")
            if idx is None:
                return result
            try:
                options_obj = s_kwargs.get("options", s_args[1] if len(s_args) > 1 else [])
                options_list = list(options_obj)
            except Exception:
                options_list = []
            if not options_list:
                return result

            row = context["captured"].setdefault(str(idx), {})
            if row.get("widget_captured"):
                return result
            fmt = s_kwargs.get("format_func")
            current_display = _safe_format(fmt, result)
            products = _self_warehouse_products(core_module, db_path)
            current_pid = _extract_product_id(result) or _pid_from_display(current_display, products)
            pid_to_value: dict[str, Any] = {}
            for opt in options_list:
                display = _safe_format(fmt, opt)
                pid = _extract_product_id(opt) or _pid_from_display(display, products)
                if pid is not None:
                    pid_to_value[str(int(pid))] = opt if _primitive(opt) else int(pid)
            row.update({
                "widget_captured": True,
                "widget_key": s_kwargs.get("key"),
                "current_pid": current_pid,
                "current_display": current_display,
                "match_rate": _match_rate(current_display),
                "pid_to_value": pid_to_value,
            })
            return result

        def section_wrapper(*sec_args, **sec_kwargs):
            result = original_section(*sec_args, **sec_kwargs)
            title = str(sec_args[0] if sec_args else sec_kwargs.get("title", ""))
            if "상품 매칭 확인" not in title or context.get("summary_rendered"):
                return result
            context["summary_rendered"] = True
            fp = context.get("file_fp")
            excel_rows = context.get("excel_rows") or []
            cache = st_obj.session_state.get(_CACHE_KEY, {})
            meta = cache.get(fp) if fp else None
            if fp and excel_rows and meta:
                st_obj.markdown(
                    """<style>
                    div[data-testid=\"stExpander\"] {display:none !important;}
                    </style>""",
                    unsafe_allow_html=True,
                )
                _render_review_table(st_obj, pd_obj, core_module, db_path, fp, excel_rows, meta)
            elif fp and excel_rows:
                st_obj.info("매칭 결과 표를 준비하고 있습니다…")
            return result

        st_obj.file_uploader = file_uploader_wrapper
        st_obj.expander = expander_wrapper
        st_obj.selectbox = selectbox_wrapper
        if callable(original_section):
            kwargs = dict(kwargs)
            kwargs["section"] = section_wrapper

        try:
            result = original_render(*args, **kwargs)
        finally:
            st_obj.file_uploader = original_file_uploader
            st_obj.expander = original_expander
            st_obj.selectbox = original_selectbox

        fp = context.get("file_fp")
        excel_rows = context.get("excel_rows") or []
        captured = context.get("captured") or {}
        if fp and excel_rows and captured:
            cache = st_obj.session_state.setdefault(_CACHE_KEY, {})
            new_meta = {
                "file_name": context.get("file_name", ""),
                "rows": captured,
            }
            had_meta = fp in cache
            cache[fp] = new_meta
            if len(cache) > 5:
                for old_key in list(cache.keys())[:-5]:
                    cache.pop(old_key, None)
            st_obj.session_state[_CACHE_KEY] = cache

            if not had_meta and st_obj.session_state.get(_READY_KEY) != fp:
                st_obj.session_state[_READY_KEY] = fp
                try:
                    st_obj.rerun()
                except Exception:
                    pass
        return result

    purchase_module.render_purchase_page = render_purchase_page
    purchase_module._rg_purchase_match_ui_v090_applied = True
    _APPLIED = True
    return purchase_module
