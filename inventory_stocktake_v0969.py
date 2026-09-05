"""RG Manager warehouse-specific stocktake workbook filtering.

Fixes v0.9.68 export where every warehouse sheet repeated the same product master.

Export eligibility:
- 자체창고: raw/self-warehouse items only.
- 쿠팡RG: finished products only.
- 반품창고: finished products only (returns are original finished goods).
- Any extra warehouse: only products that currently have non-zero stock there.
- Archived products are shown only when that specific warehouse still has stock.

v0.9.175:
- 쿠팡RG sheet pre-fills `실사수량` with the latest Coupang API
  `coupang_rg_inventory.orderable_qty` for normal stock only.
- Missing API rows remain blank instead of being treated as zero.
- ERP현재고 remains the accounting/ledger balance, so the existing difference
  formula immediately shows API saleable stock minus ERP ledger stock.

The upload/preview/commit safety rules remain those of v0.9.68.

v0.9.71 startup hook:
- ensure the three user-requested Coupang finished products exist;
- ensure same-name raw/self-warehouse counterparts exist using next-free JDS codes;
- register/update the requested BOM quantities idempotently.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import inventory_stocktake_v0968 as base


def _eligible_for_sheet(p: dict, wh_name: str, qty: float) -> bool:
    item_type = str(p.get("item_type") or "").strip().lower()
    active = bool(int(p.get("active") or 0))
    has_here = abs(float(qty or 0)) > 1e-12

    # Archived products must never clutter every worksheet.  They remain visible
    # only when physical/accounting stock still exists in this exact warehouse.
    if not active and not has_here:
        return False

    if wh_name == "자체창고":
        return item_type == "raw"
    if wh_name in ("쿠팡RG", "반품창고"):
        return item_type == "finished"

    # Unknown/auxiliary warehouses have no product-master assignment rule.
    # Export only balances that actually exist there instead of duplicating the
    # full master into another sheet.
    return has_here


def _api_rg_saleable_stock(core, db) -> dict[int, dict[str, object]]:
    """Return current normal RG API stock keyed by ERP product id.

    The current inventory table is replaced only after a successful full API
    inventory fetch, so rows in this table are the latest complete snapshot.
    Return/unmatched options are intentionally excluded from the normal RG sheet.
    """
    try:
        import coupang_api_sync_v09140 as api
        api.ensure_schema(core, db)
    except Exception:
        return {}

    try:
        with core._conn(db) as con:
            exists = con.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='coupang_rg_inventory'"
            ).fetchone()
            if not exists:
                return {}
            cols = {
                str(r["name"])
                for r in con.execute("PRAGMA table_info(coupang_rg_inventory)").fetchall()
            }
            required = {"product_id", "orderable_qty", "synced_at", "stock_type"}
            if not required.issubset(cols):
                return {}
            rows = con.execute(
                """SELECT product_id,
                          COALESCE(SUM(orderable_qty),0) qty,
                          MAX(synced_at) synced_at
                   FROM coupang_rg_inventory
                   WHERE product_id IS NOT NULL
                     AND stock_type='normal'
                   GROUP BY product_id"""
            ).fetchall()
    except Exception:
        return {}

    out: dict[int, dict[str, object]] = {}
    for r in rows:
        try:
            pid = int(r["product_id"])
            qty = float(r["qty"] or 0)
        except Exception:
            continue
        out[pid] = {
            "qty": qty,
            "synced_at": str(r["synced_at"] or "").strip(),
        }
    return out


def _build_workbook(core, db) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule
    except Exception as exc:
        raise RuntimeError(f"Excel 생성 모듈(openpyxl)을 불러오지 못했습니다: {exc}")

    products, _wh_by_id, wh_by_name, balances = base._master(core, db)
    warehouses = base._ordered_warehouses(wh_by_name)
    if not warehouses:
        raise ValueError("등록된 창고가 없습니다.")

    # Latest successfully fetched normal-product saleable inventory from Coupang.
    # This is used only as a pre-filled physical/API count in the RG worksheet;
    # the accounting ledger remains untouched until the workbook is uploaded and
    # explicitly confirmed through the existing stocktake workflow.
    api_rg_stock = _api_rg_saleable_stock(core, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "사용방법"
    ws["A1"] = "RG Manager 재고 실사"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "1. 자체창고 시트는 자체창고 품목, 쿠팡RG·반품창고 시트는 완제품만 표시됩니다."
    ws["A4"] = "2. 쿠팡RG 시트의 '실사수량'은 최근 쿠팡 API로 조회한 판매가능재고가 자동 입력됩니다."
    ws["A5"] = "3. API에 조회되지 않은 쿠팡RG 상품과 자체창고·반품창고는 실제 확인한 수량만 '실사수량'에 입력하세요."
    ws["A6"] = "4. 실사하지 않은 행은 빈칸으로 두면 조정 대상에서 제외됩니다."
    ws["A7"] = "5. ERP상품ID·창고·품목코드·쿠팡 옵션ID·상품명·ERP현재고는 수정하지 않는 것을 권장합니다."
    ws["A8"] = "6. 업로드 시 ERP의 현재 재고를 다시 조회한 뒤 차이만 '재고실사조정' 이력으로 반영합니다."
    ws["A9"] = "7. 같은 파일은 중복 적용할 수 없습니다."
    ws["A11"] = "다운로드 시각"
    ws["B11"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.column_dimensions["A"].width = 110
    ws.column_dimensions["B"].width = 22

    header_fill = PatternFill("solid", fgColor="CFE3FF")
    header_font = Font(bold=True, color="0D3768")
    thin = Side(style="thin", color="D8E0EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    plus_fill = PatternFill("solid", fgColor="E2F0D9")
    minus_fill = PatternFill("solid", fgColor="FCE4D6")

    used = {"사용방법"}
    headers = [
        "ERP상품ID", "창고", "품목코드", "쿠팡 옵션ID", "상품명", "상태",
        "ERP현재고", "실사수량", "차이", "비고",
    ]

    for wh_name in warehouses:
        wid = wh_by_name[wh_name]
        sh = wb.create_sheet(base._sheet_name(wh_name, used))
        sh.append(headers)
        for cell in sh[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        rows = []
        for pid, p in products.items():
            qty = balances.get((pid, wid), 0.0)
            if not _eligible_for_sheet(p, wh_name, qty):
                continue

            actual_qty = None
            note = None
            if wh_name == "쿠팡RG" and pid in api_rg_stock:
                api_row = api_rg_stock[pid]
                actual_qty = float(api_row.get("qty") or 0)
                synced_at = str(api_row.get("synced_at") or "").strip()
                note = "쿠팡 API 판매가능재고 자동입력"
                if synced_at:
                    note += f" · 조회 {synced_at}"

            rows.append((
                p["name"].lower(), p["display_code"], pid,
                [pid, wh_name, p["display_code"], p["option_id"], p["name"],
                 "사용중" if p["active"] else "보관", qty, actual_qty, None, note],
            ))

        rows.sort(key=lambda x: (x[0], x[1], x[2]))
        for _name, _code, _pid, values in rows:
            sh.append(values)

        last = max(2, sh.max_row)
        if sh.max_row < 2:
            # Keep one blank-looking protected/formatted row only for worksheet
            # structure; no product ID means upload parser ignores it.
            sh.append([None, wh_name, None, None, "표시할 품목이 없습니다.", None, None, None, None, None])
            last = sh.max_row

        for r in range(2, last + 1):
            if sh.cell(r, 1).value is not None:
                sh.cell(r, 9).value = f'=IF(H{r}="","",H{r}-G{r})'
                sh.cell(r, 8).fill = editable_fill
            for c in range(1, 11):
                sh.cell(r, c).border = border
                sh.cell(r, c).alignment = Alignment(vertical="center")
            for c in (7, 8, 9):
                sh.cell(r, c).number_format = '#,##0.##'

        dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        dv.error = "실사수량은 0 이상의 숫자로 입력하세요."
        dv.errorTitle = "실사수량 입력 오류"
        sh.add_data_validation(dv)
        data_last = max(2, sh.max_row)
        dv.add(f"H2:H{data_last}")
        sh.conditional_formatting.add(
            f"I2:I{data_last}", CellIsRule(operator="greaterThan", formula=["0"], fill=plus_fill)
        )
        sh.conditional_formatting.add(
            f"I2:I{data_last}", CellIsRule(operator="lessThan", formula=["0"], fill=minus_fill)
        )

        sh.freeze_panes = "A2"
        sh.auto_filter.ref = f"A1:J{data_last}"
        sh.column_dimensions["A"].hidden = True
        widths = {"B": 14, "C": 18, "D": 18, "E": 52, "F": 10, "G": 14, "H": 14, "I": 14, "J": 40}
        for col, width in widths.items():
            sh.column_dimensions[col].width = width
        sh.row_dimensions[1].height = 24

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def apply(core_module):
    # v0.9.71: create/reactivate requested product masters and set their BOMs.
    # The seed is idempotent; Streamlit reruns never duplicate products/BOM rows.
    try:
        import requested_product_bom_seed_v0971 as seed
        seed.apply(core_module)
    except Exception as exc:
        # Keep the ERP available even if a local legacy schema causes a seed error.
        setattr(core_module, "_rg_requested_product_bom_seed_v0971_error", str(exc))

    # Patch the v0.9.68 exporter before its existing UI wrapper is installed.
    base._build_workbook = _build_workbook
    return base.apply(core_module)
