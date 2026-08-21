"""RG Manager v0.9.125: manual match or discard for unmatched sales rows.

When a Coupang sales-stat workbook contains option IDs that cannot be safely
matched automatically, no sales data is written yet. The user can resolve each
row by either:
- searching an existing ERP finished product and manually mapping the option, or
- discarding that option from this upload.

Manual matches are persisted in return_discount_aliases so the same Coupang
option ID is automatically linked to the same ERP product in future uploads.
Discarded rows are physically removed from an in-memory workbook copy before the
existing sales import pipeline runs.
"""
from __future__ import annotations

from io import BytesIO
import re
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

_APPLIED = False
_CORE = None
_RD = None
_PREVIOUS_IMPORT = None
_DEFAULT_DB = None

_PENDING_KEY = "_rg_v09125_unmatched_sales_pending"
_FLASH_KEY = "_rg_v09125_unmatched_sales_flash"


def _oid(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    try:
        x = float(s)
        if x == x and abs(x - round(x)) < 1e-9:
            return str(int(round(x)))
    except Exception:
        pass
    return s[:-2] if s.endswith(".0") and s[:-2].isdigit() else s


def _source_bytes(source) -> bytes | None:
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if hasattr(source, "getvalue"):
        try:
            return bytes(source.getvalue())
        except Exception:
            pass
    if hasattr(source, "read"):
        try:
            pos = source.tell() if hasattr(source, "tell") else None
            data = source.read()
            if pos is not None and hasattr(source, "seek"):
                source.seek(pos)
            return bytes(data)
        except Exception:
            pass
    return None


def _parse_unmatched_error(exc: Any, parsed: list[dict] | None = None) -> list[dict]:
    text = str(exc or "")
    signals = (
        "원상품을 안전하게 자동 매칭할 수 없습니다",
        "ERP에 없는 쿠팡 옵션ID",
        "품목관리에 없는 판매 옵션",
    )
    if not any(x in text for x in signals):
        return []

    parsed_by_oid = {
        _oid(r.get("option_id")): r
        for r in (parsed or [])
        if _oid(r.get("option_id"))
    }
    out = []
    for raw in text.splitlines():
        line = raw.strip()
        if "|" not in line:
            continue
        left, right = line.split("|", 1)
        oid = _oid(left)
        if not oid.isdigit():
            continue
        label = right.strip()
        reason = "원상품 자동 매칭 불가"
        if label.endswith(")") and " (" in label:
            label, tail = label.rsplit(" (", 1)
            reason = tail[:-1].strip() or reason

        src = parsed_by_oid.get(oid, {})
        qty = src.get("qty")
        try:
            qty = float(qty) if qty is not None else None
        except Exception:
            qty = None
        out.append(
            {
                "option_id": oid,
                "name": str(src.get("name") or label).strip(),
                "qty": qty,
                "reason": reason,
            }
        )
    return out


def _merge_rows(old, new):
    merged = {
        _oid(r.get("option_id")): dict(r)
        for r in old
        if _oid(r.get("option_id"))
    }
    for r in new:
        oid = _oid(r.get("option_id"))
        if oid:
            merged[oid] = dict(r)
    return list(merged.values())


def _header_key(value: Any) -> str:
    return re.sub(r"[\s_()\-]+", "", str(value or "")).lower()


def _filtered_workbook(raw: bytes, ignored_ids: set[str]) -> tuple[bytes, int]:
    """Return an xlsx copy with only user-discarded option rows removed."""
    if not raw:
        raise ValueError("업로드한 판매통계 파일을 다시 읽지 못했습니다.")
    if not ignored_ids:
        return raw, 0

    try:
        wb = load_workbook(BytesIO(raw))
    except Exception as exc:
        raise ValueError(f"판매통계 Excel을 다시 열지 못했습니다: {exc}") from exc

    ws = wb["판매통계"] if "판매통계" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = None
    option_col = None

    scan_rows = min(max(int(ws.max_row or 1), 1), 15)
    for r in range(1, scan_rows + 1):
        for c in range(1, int(ws.max_column or 1) + 1):
            key = _header_key(ws.cell(r, c).value)
            if key == "옵션id" or key.endswith("옵션id"):
                header_row, option_col = r, c
                break
        if option_col is not None:
            break

    if header_row is None or option_col is None:
        raise ValueError("판매통계 시트에서 옵션ID 열을 찾지 못했습니다.")

    removed = 0
    for r in range(int(ws.max_row or 0), header_row, -1):
        if _oid(ws.cell(r, option_col).value) in ignored_ids:
            ws.delete_rows(r, 1)
            removed += 1

    if removed <= 0:
        raise ValueError(
            "버리기로 선택한 옵션ID가 판매통계 시트에서 발견되지 않았습니다."
        )

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), removed


def _pending(source, file_name, period_start, period_end, db_path, rows):
    raw = _source_bytes(source)
    if not raw:
        raise ValueError("업로드한 판매통계 파일을 다시 읽지 못했습니다.")
    return {
        "source": raw,
        "file_name": str(file_name or ""),
        "period_start": str(period_start or ""),
        "period_end": str(period_end or ""),
        "db_path": str(db_path or _DEFAULT_DB),
        "rows": rows,
        "decisions": {},
    }


def _load_match_products(core, rd, db):
    """Load user-facing ERP finished products, including archived historical items."""
    core.init_db(db)
    hidden = set()
    try:
        import product_visibility_v0995 as visibility
        hidden = set(visibility.hidden_ids(core, db))
    except Exception:
        hidden = set()

    with core._conn(db) as c:
        rows = c.execute(
            """SELECT id,item_code,option_id,name,item_type,unit_cost,active
               FROM products
               WHERE item_type='finished'
               ORDER BY active DESC,name,item_code"""
        ).fetchall()

    out = []
    for r in rows:
        pid = int(r["id"])
        if pid in hidden:
            continue
        p = {
            "id": pid,
            "item_code": str(r["item_code"] or ""),
            "option_id": _oid(r["option_id"]),
            "name": str(r["name"] or ""),
            "item_type": str(r["item_type"] or ""),
            "unit_cost": float(r["unit_cost"] or 0),
            "active": int(r["active"] or 0),
        }
        try:
            if rd._placeholder(p):
                continue
        except Exception:
            pass
        if not p["option_id"]:
            continue
        out.append(p)
    return out


def _product_label(p: dict) -> str:
    state = "" if int(p.get("active") or 0) else " · 보관상품"
    return (
        f"{p.get('name','')} · 옵션ID {p.get('option_id','')}"
        f"{state}"
    )


def _search_products(products: list[dict], query: str, limit: int = 30) -> list[dict]:
    q = re.sub(r"\s+", " ", str(query or "").strip().lower())
    if not q:
        return []
    terms = [x for x in q.split(" ") if x]

    def score(p):
        name = str(p.get("name") or "").lower()
        oid = str(p.get("option_id") or "").lower()
        code = str(p.get("item_code") or "").lower()
        hay = f"{name} {oid} {code}"
        if not all(t in hay for t in terms):
            return None
        s = 0
        if q == oid or q == code:
            s += 100
        if q in name:
            s += 40
        if name.startswith(q):
            s += 20
        if int(p.get("active") or 0):
            s += 5
        return s

    ranked = []
    for p in products:
        s = score(p)
        if s is not None:
            ranked.append((s, p))
    ranked.sort(key=lambda x: (-x[0], str(x[1].get("name") or "")))
    return [p for _s, p in ranked[:limit]]


def _snapshot_aliases(core, db, option_ids: set[str]) -> dict[str, dict]:
    if not option_ids:
        return {}
    marks = ",".join("?" for _ in option_ids)
    with core._conn(db) as c:
        rows = c.execute(
            f"""SELECT discount_option_id,parent_product_id,discount_name,
                       match_method,created_at,updated_at
                FROM return_discount_aliases
                WHERE discount_option_id IN ({marks})""",
            tuple(sorted(option_ids)),
        ).fetchall()
    return {str(r["discount_option_id"]): dict(r) for r in rows}


def _install_manual_aliases(core, rd, db, manual_map: dict[str, int], rows: list[dict]):
    if not manual_map:
        return {}
    rd._ensure_schema(core, db)
    ids = set(manual_map)
    previous = _snapshot_aliases(core, db, ids)
    by_oid = {_oid(r.get("option_id")): r for r in rows}
    now = core.now_iso()
    with core._conn(db) as c:
        for oid, parent_pid in manual_map.items():
            row = by_oid.get(oid, {})
            c.execute(
                """INSERT INTO return_discount_aliases
                   (discount_option_id,parent_product_id,discount_name,
                    match_method,created_at,updated_at)
                   VALUES(?,?,?,?,?,?)
                   ON CONFLICT(discount_option_id) DO UPDATE SET
                     parent_product_id=excluded.parent_product_id,
                     discount_name=excluded.discount_name,
                     match_method=excluded.match_method,
                     updated_at=excluded.updated_at""",
                (
                    oid,
                    int(parent_pid),
                    str(row.get("name") or ""),
                    "manual_user",
                    now,
                    now,
                ),
            )
    return previous


def _restore_aliases(core, db, option_ids: set[str], previous: dict[str, dict]):
    if not option_ids:
        return
    marks = ",".join("?" for _ in option_ids)
    with core._conn(db) as c:
        c.execute(
            f"DELETE FROM return_discount_aliases WHERE discount_option_id IN ({marks})",
            tuple(sorted(option_ids)),
        )
        for oid, row in previous.items():
            c.execute(
                """INSERT INTO return_discount_aliases
                   (discount_option_id,parent_product_id,discount_name,
                    match_method,created_at,updated_at)
                   VALUES(?,?,?,?,?,?)""",
                (
                    oid,
                    int(row["parent_product_id"]),
                    str(row.get("discount_name") or ""),
                    str(row.get("match_method") or "manual_user"),
                    str(row.get("created_at") or core.now_iso()),
                    str(row.get("updated_at") or core.now_iso()),
                ),
            )


def _dialog_body():
    pending = st.session_state.get(_PENDING_KEY)
    if not pending:
        st.info("확인할 미매칭 판매 데이터가 없습니다.")
        return

    rows = list(pending.get("rows") or [])
    decisions = pending.setdefault("decisions", {})
    db = pending.get("db_path") or _DEFAULT_DB

    try:
        products = _load_match_products(_CORE, _RD, db)
    except Exception as exc:
        st.error(f"ERP 상품목록을 불러오지 못했습니다. {exc}")
        return
    product_by_id = {int(p["id"]): p for p in products}

    st.warning(
        f"품목관리에 안전하게 매칭되지 않는 판매 옵션 {len(rows)}개가 있습니다."
    )
    st.write(
        "각 항목마다 **ERP 상품에 직접 매칭**하거나 **이번 자료에서 버리기**를 선택하세요. "
        "직접 매칭한 옵션ID는 다음 판매통계부터 같은 ERP 상품으로 자동 연결됩니다."
    )

    for i, row in enumerate(rows, start=1):
        oid = _oid(row.get("option_id"))
        decision = decisions.get(oid)

        with st.container(border=True):
            st.markdown(f"**{i}. {row.get('name','')}**")
            st.caption(
                f"옵션ID {oid} · 판매수량 {row.get('qty','')} · "
                f"{row.get('reason','원상품 자동 매칭 불가')}"
            )

            if decision:
                if decision.get("action") == "discard":
                    st.warning("현재 선택: **이 항목 버리기**")
                elif decision.get("action") == "match":
                    try:
                        p = product_by_id.get(int(decision.get("product_id")))
                    except Exception:
                        p = None
                    if p:
                        st.success(f"현재 매칭: **{_product_label(p)}**")
                    else:
                        decisions.pop(oid, None)
                        decision = None

            q = st.text_input(
                "ERP 상품 검색",
                key=f"_rg_v09125_search_{oid}",
                placeholder="상품명 또는 옵션ID를 입력하세요",
            )
            hits = _search_products(products, q)
            selected_pid = None
            if q:
                if hits:
                    options = [int(p["id"]) for p in hits]
                    selected_pid = st.selectbox(
                        "검색 결과",
                        options=options,
                        format_func=lambda pid: _product_label(product_by_id[int(pid)]),
                        key=f"_rg_v09125_result_{oid}",
                    )
                else:
                    st.caption("검색 결과가 없습니다.")

            c1, c2, c3 = st.columns([1.35, 1, 1])
            if c1.button(
                "선택 상품으로 매칭",
                key=f"_rg_v09125_match_{oid}",
                use_container_width=True,
                disabled=selected_pid is None,
            ):
                decisions[oid] = {
                    "action": "match",
                    "product_id": int(selected_pid),
                }
                pending["decisions"] = decisions
                st.session_state[_PENDING_KEY] = pending
                st.success(
                    f"매칭 선택됨: {_product_label(product_by_id[int(selected_pid)])}"
                )

            if c2.button(
                "이 항목 버리기",
                key=f"_rg_v09125_discard_{oid}",
                use_container_width=True,
            ):
                decisions[oid] = {"action": "discard"}
                pending["decisions"] = decisions
                st.session_state[_PENDING_KEY] = pending
                st.warning("버리기로 선택했습니다.")

            if c3.button(
                "선택 취소",
                key=f"_rg_v09125_clear_{oid}",
                use_container_width=True,
                disabled=oid not in decisions,
            ):
                decisions.pop(oid, None)
                pending["decisions"] = decisions
                st.session_state[_PENDING_KEY] = pending
                st.info("선택을 취소했습니다.")

    decisions = pending.get("decisions", {})
    resolved = sum(
        1
        for r in rows
        if decisions.get(_oid(r.get("option_id")), {}).get("action")
        in {"match", "discard"}
    )
    st.markdown(f"**처리 선택 {resolved}/{len(rows)}개 완료**")

    c1, c2 = st.columns([1.4, 1])
    submit = c1.button(
        "선택대로 판매통계 입력",
        type="primary",
        use_container_width=True,
        disabled=resolved != len(rows),
        key="_rg_v09125_submit",
    )
    cancel = c2.button(
        "전체 취소",
        use_container_width=True,
        key="_rg_v09125_cancel",
    )

    if cancel:
        st.session_state.pop(_PENDING_KEY, None)
        st.rerun()

    if not submit:
        if resolved != len(rows):
            st.caption("모든 미매칭 항목을 매칭 또는 버리기로 선택하면 입력할 수 있습니다.")
        return

    discard_ids = {
        oid for oid, d in decisions.items() if d.get("action") == "discard"
    }
    manual_map = {
        oid: int(d["product_id"])
        for oid, d in decisions.items()
        if d.get("action") == "match" and d.get("product_id") is not None
    }

    try:
        filtered, removed_rows = _filtered_workbook(
            bytes(pending["source"]), discard_ids
        )
    except Exception as exc:
        st.error(f"입력용 판매통계 파일을 만들지 못했습니다. {exc}")
        return

    previous_aliases = {}
    if manual_map:
        try:
            previous_aliases = _install_manual_aliases(
                _CORE, _RD, db, manual_map, rows
            )
        except Exception as exc:
            st.error(f"수동 매칭을 저장하지 못했습니다. {exc}")
            return

    source = BytesIO(filtered)
    try:
        source.name = pending.get("file_name") or "sales.xlsx"
    except Exception:
        pass

    try:
        _PREVIOUS_IMPORT(
            source,
            pending.get("file_name", ""),
            pending.get("period_start", ""),
            pending.get("period_end", ""),
            db,
        )
    except ValueError as exc:
        if manual_map:
            try:
                _restore_aliases(_CORE, db, set(manual_map), previous_aliases)
            except Exception:
                pass
        try:
            parsed = _RD._parse_sales_file(filtered)
        except Exception:
            parsed = []
        extra = _parse_unmatched_error(exc, parsed)
        if extra:
            pending["rows"] = _merge_rows(rows, extra)
            st.session_state[_PENDING_KEY] = pending
            st.error(
                "추가로 매칭되지 않는 옵션이 있습니다. 새로 표시된 항목도 처리해 주세요."
            )
            return
        st.error(f"자료를 반영하지 못했습니다. {exc}")
        return
    except Exception as exc:
        if manual_map:
            try:
                _restore_aliases(_CORE, db, set(manual_map), previous_aliases)
            except Exception:
                pass
        st.error(f"자료를 반영하지 못했습니다. {exc}")
        return

    st.session_state.pop(_PENDING_KEY, None)
    parts = []
    if manual_map:
        parts.append(f"{len(manual_map)}개 수동 매칭")
    if discard_ids:
        parts.append(f"{len(discard_ids)}개 버림({removed_rows}행)")
    summary = " · ".join(parts) if parts else "미매칭 처리 완료"
    st.session_state[_FLASH_KEY] = f"{summary} 후 나머지 판매통계를 반영했습니다."
    st.rerun()


def _make_dialog():
    dialog = getattr(st, "dialog", None)
    if not callable(dialog):
        return _dialog_body
    try:
        return dialog("매칭되지 않는 판매 데이터 확인", width="large")(_dialog_body)
    except TypeError:
        return dialog("매칭되지 않는 판매 데이터 확인")(_dialog_body)


_dialog = _make_dialog()


def apply(core, return_discount_module, db_path=None):
    global _APPLIED, _CORE, _RD, _PREVIOUS_IMPORT, _DEFAULT_DB

    _CORE = core
    _RD = return_discount_module
    _DEFAULT_DB = db_path or core.DEFAULT_DB

    flash = st.session_state.pop(_FLASH_KEY, None)
    if flash:
        try:
            st.toast(str(flash), icon="✅")
        except Exception:
            st.success(str(flash))

    if _APPLIED or getattr(
        core, "_rg_sales_unmatched_confirm_v09125_applied", False
    ):
        return core

    previous_import = core.import_sales_stats
    _PREVIOUS_IMPORT = previous_import

    def import_sales_stats(
        source, file_name, period_start, period_end, db_path=None
    ):
        target = db_path or _DEFAULT_DB

        # Preflight only. No sales/inventory/P&L data is written before the user
        # resolves every unmatched row.
        try:
            parsed = return_discount_module._parse_sales_file(source)
        except Exception:
            parsed = []

        if parsed:
            try:
                return_discount_module._resolve(core, target, parsed)
            except ValueError as exc:
                rows = _parse_unmatched_error(exc, parsed)
                if rows:
                    st.session_state[_PENDING_KEY] = _pending(
                        source,
                        file_name,
                        period_start,
                        period_end,
                        target,
                        rows,
                    )
                    _dialog()
                    st.stop()

        try:
            return previous_import(
                source, file_name, period_start, period_end, target
            )
        except ValueError as exc:
            rows = _parse_unmatched_error(exc, parsed)
            if not rows:
                raise
            st.session_state[_PENDING_KEY] = _pending(
                source,
                file_name,
                period_start,
                period_end,
                target,
                rows,
            )
            _dialog()
            st.stop()

    core.import_sales_stats = import_sales_stats
    core._rg_sales_unmatched_confirm_v09125_applied = True
    _APPLIED = True
    return core
