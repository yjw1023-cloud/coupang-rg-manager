"""RG Manager v0.9.5 batch production from Coupang Rocket Growth inbound Excel.

Rules
- Production targets are rows with a value in column V (입고 수량 입력) of sheet
  `로켓그로스 입고`.
- Product matching uses Coupang option ID from column G first.
- Before ANY inventory/production write, every target row must be valid and every
  matched finished product must have a BOM.
- If even one target has no BOM, the whole batch is blocked. No partial production.
- Actual batch execution is a single SQLite transaction: 자체창고 BOM consumption
  -> 쿠팡RG finished-goods receipt, preserving the current negative-stock rule.
- File hash is unique after successful execution so the same production file cannot
  be executed twice.
"""
from __future__ import annotations

from datetime import datetime
import hashlib
import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_NAME = "로켓그로스 입고"
START_ROW = 5
MAX_QTY = 5000


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _display_code(item_code: Any, option_id: Any = None) -> str:
    code = _text(item_code)
    if re.fullmatch(r"CP-\d+", code):
        return _text(option_id) or code[3:]
    return code


def _file_bytes(uploaded: Any) -> bytes:
    if uploaded is None:
        return b""
    if isinstance(uploaded, (str, Path)):
        return Path(uploaded).read_bytes()
    if isinstance(uploaded, (bytes, bytearray)):
        return bytes(uploaded)
    if hasattr(uploaded, "getvalue"):
        try:
            return bytes(uploaded.getvalue())
        except Exception:
            pass
    pos = None
    try:
        pos = uploaded.tell()
    except Exception:
        pass
    try:
        data = uploaded.read()
        return data.encode("utf-8") if isinstance(data, str) else bytes(data or b"")
    finally:
        if pos is not None:
            try:
                uploaded.seek(pos)
            except Exception:
                pass


def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _parse_qty(value: Any) -> tuple[int | None, str]:
    if value is None or _text(value) == "":
        return None, ""
    try:
        n = float(value)
    except Exception:
        return None, "생산수량이 숫자가 아닙니다."
    if not n.is_integer():
        return None, "생산수량은 정수여야 합니다."
    q = int(n)
    if q < 1 or q > MAX_QTY:
        return None, f"생산수량은 1~{MAX_QTY:,} 사이여야 합니다."
    return q, ""


def parse_production_excel(source: Any) -> dict[str, Any]:
    """Parse only explicitly entered production rows. No DB write is performed."""
    data = _file_bytes(source)
    if not data:
        raise ValueError("생산자료 Excel 파일이 비어 있습니다.")

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"'{SHEET_NAME}' 시트를 찾지 못했습니다. 쿠팡 로켓그로스 입고 양식을 사용해 주세요.")
        ws = wb[SHEET_NAME]

        option_header = _text(ws["G3"].value).replace(" ", "")
        qty_header = _text(ws["V3"].value).replace(" ", "").replace("\n", "")
        if "옵션ID" not in option_header or "입고수량입력" not in qty_header:
            raise ValueError("생산자료 양식이 다릅니다. G열 옵션 ID / V열 입고 수량 입력 양식을 확인해 주세요.")

        rows: list[dict[str, Any]] = []
        for r in range(START_ROW, ws.max_row + 1):
            raw_qty = ws.cell(r, 22).value
            if raw_qty is None or _text(raw_qty) == "":
                continue
            qty, qty_error = _parse_qty(raw_qty)
            rows.append({
                "source_row": r,
                "source_name": _text(ws.cell(r, 2).value),
                "option_name": _text(ws.cell(r, 3).value),
                "option_id": _text(ws.cell(r, 7).value),
                "qty": qty,
                "raw_qty": _text(raw_qty),
                "parse_error": qty_error,
            })
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not rows:
        raise ValueError("V열 '입고 수량 입력'에 생산수량이 입력된 상품이 없습니다.")
    return {"data": data, "file_hash": file_hash(data), "rows": rows}


def ensure_schema(core_module, db_path=None) -> None:
    db_path = db_path or core_module.DEFAULT_DB
    core_module.init_db(db_path)
    with core_module._conn(db_path) as c:
        c.executescript(
            """
            CREATE TABLE IF NOT EXISTS production_batch_imports(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_hash TEXT NOT NULL UNIQUE,
                file_name TEXT,
                production_date TEXT NOT NULL,
                target_rows INTEGER NOT NULL,
                total_qty REAL NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS production_batch_lines(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                source_row INTEGER,
                option_id TEXT,
                product_id INTEGER NOT NULL,
                qty REAL NOT NULL,
                produced_unit_cost REAL NOT NULL DEFAULT 0,
                ref_no TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(batch_id) REFERENCES production_batch_imports(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );
            CREATE INDEX IF NOT EXISTS ix_production_batch_lines_batch ON production_batch_lines(batch_id);
            CREATE INDEX IF NOT EXISTS ix_production_batch_lines_product ON production_batch_lines(product_id);
            """
        )


def _find_product(c, option_id: str):
    option_id = _text(option_id)
    if not option_id:
        return None, "옵션ID가 없습니다."
    rows = c.execute(
        "SELECT id,item_code,option_id,name FROM products WHERE CAST(option_id AS TEXT)=?",
        (option_id,),
    ).fetchall()
    if not rows:
        rows = c.execute(
            "SELECT id,item_code,option_id,name FROM products WHERE item_code IN (?,?)",
            (option_id, f"CP-{option_id}"),
        ).fetchall()
    if len(rows) == 1:
        return rows[0], ""
    if len(rows) > 1:
        return None, "ERP에 같은 옵션ID로 연결 가능한 상품이 여러 개 있습니다."
    return None, "ERP에 해당 옵션ID 상품이 등록되어 있지 않습니다."


def validate_rows(core_module, parsed_rows: list[dict[str, Any]], db_path=None) -> dict[str, Any]:
    """Validate the complete batch before production. Never writes inventory."""
    db_path = db_path or core_module.DEFAULT_DB
    ensure_schema(core_module, db_path)
    results: list[dict[str, Any]] = []
    seen: dict[str, int] = {}

    with core_module._conn(db_path) as c:
        for src in parsed_rows:
            r = dict(src)
            errors: list[str] = []
            if r.get("parse_error"):
                errors.append(str(r["parse_error"]))
            option_id = _text(r.get("option_id"))
            if not option_id:
                errors.append("옵션ID가 비어 있습니다.")
            if option_id:
                if option_id in seen:
                    errors.append(f"같은 파일에 옵션ID가 중복 입력되었습니다. (첫 행 {seen[option_id]})")
                else:
                    seen[option_id] = int(r.get("source_row") or 0)

            product = None
            product_error = ""
            if option_id:
                product, product_error = _find_product(c, option_id)
            if product_error:
                errors.append(product_error)

            r.update({
                "product_id": None,
                "item_code": "",
                "erp_name": "",
                "bom_count": 0,
                "unit_cost": 0.0,
                "missing_bom": False,
            })

            if product is not None:
                pid = int(product["id"])
                bom = c.execute(
                    """SELECT b.component_product_id,b.qty_per,p.unit_cost
                       FROM bom_items b
                       JOIN products p ON p.id=b.component_product_id
                       WHERE b.parent_product_id=?""",
                    (pid,),
                ).fetchall()
                if not bom:
                    errors.append("BOM이 없습니다.")
                    r["missing_bom"] = True
                elif any(int(x["component_product_id"]) == pid for x in bom):
                    errors.append("BOM 오류: 완제품이 자기 자신의 구성품입니다.")
                r.update({
                    "product_id": pid,
                    "item_code": _display_code(product["item_code"], product["option_id"]),
                    "erp_name": _text(product["name"]),
                    "bom_count": len(bom),
                    "unit_cost": sum(float(x["qty_per"] or 0) * float(x["unit_cost"] or 0) for x in bom),
                })

            r["errors"] = errors
            r["status"] = "정상" if not errors else " / ".join(errors)
            results.append(r)

    missing_bom = [r for r in results if r.get("missing_bom")]
    blocked = [r for r in results if r.get("errors")]
    return {
        "rows": results,
        "ok": not blocked,
        "blocked": blocked,
        "missing_bom": missing_bom,
        "total_qty": sum(int(r.get("qty") or 0) for r in results if r.get("qty")),
    }


def _already_executed(core_module, hash_value: str, db_path=None):
    db_path = db_path or core_module.DEFAULT_DB
    ensure_schema(core_module, db_path)
    with core_module._conn(db_path) as c:
        return c.execute(
            "SELECT id,file_name,production_date,created_at FROM production_batch_imports WHERE file_hash=?",
            (hash_value,),
        ).fetchone()


def execute_batch(core_module, parsed: dict[str, Any], file_name: str, production_date: Any, db_path=None) -> dict[str, Any]:
    """Execute all rows atomically after revalidation. Any error rolls back all rows."""
    db_path = db_path or core_module.DEFAULT_DB
    ensure_schema(core_module, db_path)
    validation = validate_rows(core_module, parsed["rows"], db_path)
    if not validation["ok"]:
        if validation["missing_bom"]:
            names = ", ".join(_text(r.get("source_name")) or _text(r.get("option_id")) for r in validation["missing_bom"][:8])
            raise ValueError(f"BOM이 없는 상품이 있습니다. 전체 생산을 중단합니다: {names}")
        raise ValueError("생산자료에 오류가 있어 전체 생산을 중단합니다. 오류 행을 확인해 주세요.")

    hash_value = _text(parsed.get("file_hash")) or file_hash(parsed.get("data") or b"")
    prod_date = core_module.norm_date(production_date)
    now = core_module.now_iso()
    ref_no = f"PRODBATCH-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"

    with core_module._conn(db_path) as c:
        c.execute("BEGIN IMMEDIATE")
        try:
            dup = c.execute(
                "SELECT id,file_name,production_date FROM production_batch_imports WHERE file_hash=?",
                (hash_value,),
            ).fetchone()
            if dup:
                raise ValueError(
                    f"이미 생산 처리한 동일 파일입니다. 기존 생산일: {dup['production_date']} / 파일: {dup['file_name'] or '-'}"
                )

            own = c.execute("SELECT id FROM warehouses WHERE name='자체창고'").fetchone()
            rg = c.execute("SELECT id FROM warehouses WHERE name='쿠팡RG'").fetchone()
            if not own or not rg:
                raise ValueError("자체창고 또는 쿠팡RG 창고를 찾지 못했습니다.")
            own_id, rg_id = int(own["id"]), int(rg["id"])

            prepared = []
            for r in validation["rows"]:
                pid = int(r["product_id"])
                bom = c.execute(
                    """SELECT b.component_product_id,b.qty_per,p.unit_cost
                       FROM bom_items b JOIN products p ON p.id=b.component_product_id
                       WHERE b.parent_product_id=?""",
                    (pid,),
                ).fetchall()
                if not bom:
                    raise ValueError(f"BOM이 없는 상품이 있습니다. 전체 생산을 중단합니다: {r['erp_name'] or r['source_name']}")
                if any(int(x["component_product_id"]) == pid for x in bom):
                    raise ValueError(f"BOM 오류가 있어 전체 생산을 중단합니다: {r['erp_name'] or r['source_name']}")
                unit_cost = sum(float(x["qty_per"] or 0) * float(x["unit_cost"] or 0) for x in bom)
                prepared.append((r, bom, unit_cost))

            cur = c.execute(
                """INSERT INTO production_batch_imports
                   (file_hash,file_name,production_date,target_rows,total_qty,created_at)
                   VALUES (?,?,?,?,?,?)""",
                (hash_value, _text(file_name), prod_date, len(prepared), validation["total_qty"], now),
            )
            batch_id = int(cur.lastrowid)

            for r, bom, unit_cost in prepared:
                pid = int(r["product_id"])
                qty = int(r["qty"])
                memo = f"생산자료 Excel {_text(file_name)} / 옵션ID {_text(r['option_id'])}"

                for b in bom:
                    need = float(b["qty_per"] or 0) * qty
                    c.execute(
                        """INSERT INTO inventory_txns
                           (txn_date,product_id,warehouse_id,qty_delta,txn_type,ref_no,memo,created_at)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (prod_date, int(b["component_product_id"]), own_id, -need,
                         "생산소모", ref_no, memo, now),
                    )

                c.execute(
                    """INSERT INTO inventory_txns
                       (txn_date,product_id,warehouse_id,qty_delta,txn_type,unit_cost,ref_no,memo,created_at)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (prod_date, pid, rg_id, qty, "생산RG입고", unit_cost, ref_no, memo, now),
                )
                c.execute(
                    "UPDATE products SET unit_cost=?,updated_at=? WHERE id=?",
                    (unit_cost, now, pid),
                )
                c.execute(
                    """INSERT INTO production_orders
                       (production_date,parent_product_id,qty,warehouse_id,produced_unit_cost,memo,created_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (prod_date, pid, qty, rg_id, unit_cost, memo, now),
                )
                c.execute(
                    """INSERT INTO production_batch_lines
                       (batch_id,source_row,option_id,product_id,qty,produced_unit_cost,ref_no,created_at)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (batch_id, int(r["source_row"]), _text(r["option_id"]), pid, qty, unit_cost, ref_no, now),
                )

            c.commit()
            return {
                "status": "produced",
                "batch_id": batch_id,
                "rows": len(prepared),
                "total_qty": validation["total_qty"],
                "ref_no": ref_no,
            }
        except Exception:
            c.rollback()
            raise


def _preview_frame(pd_obj, rows: list[dict[str, Any]]):
    return pd_obj.DataFrame([
        {
            "원본행": int(r.get("source_row") or 0),
            "상품명": _text(r.get("source_name")),
            "옵션명": _text(r.get("option_name")),
            "옵션ID": _text(r.get("option_id")),
            "생산수량": int(r.get("qty") or 0) if r.get("qty") else _text(r.get("raw_qty")),
            "ERP 품목코드": _text(r.get("item_code")),
            "ERP 상품명": _text(r.get("erp_name")),
            "BOM": f"{int(r.get('bom_count') or 0)}개 구성품" if r.get("bom_count") else "없음",
            "상태": _text(r.get("status")),
        }
        for r in rows
    ])


def render_production_batch_page(st, pd, date, core, page_header, section, **_kwargs):
    db_path = core.DEFAULT_DB
    ensure_schema(core, db_path)
    page_header(
        "생산자료",
        "쿠팡 로켓그로스 입고 Excel의 '입고 수량'을 생산수량으로 읽어 일괄 생산합니다.",
        eyebrow="PRODUCTION BATCH",
    )
    st.info(
        "파일을 업로드하는 것만으로는 생산되지 않습니다. 먼저 전체 상품의 옵션ID와 BOM을 검사하고, "
        "모든 상품이 정상일 때만 마지막 생산 실행 버튼이 활성화됩니다."
    )
    st.caption("양식 기준: 시트 '로켓그로스 입고' · G열 옵션 ID · V열 입고 수량 입력. V열이 빈 상품은 생산대상에서 제외됩니다.")

    uploaded = st.file_uploader(
        "생산자료 Excel",
        type=["xlsx"],
        key="production_batch_v095_upload",
        help="쿠팡 로켓그로스 입고요청 Excel 양식을 그대로 사용합니다.",
    )
    production_date = st.date_input("생산일", value=date.today(), key="production_batch_v095_date")

    if uploaded is None:
        return

    try:
        parsed = parse_production_excel(uploaded)
        validation = validate_rows(core, parsed["rows"], db_path)
    except Exception as exc:
        st.error(f"생산자료 확인 실패: {exc}")
        return

    duplicate = _already_executed(core, parsed["file_hash"], db_path)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("생산대상", f"{len(validation['rows']):,}개 상품")
    c2.metric("총 생산수량", f"{validation['total_qty']:,}개")
    c3.metric("BOM 정상", f"{len(validation['rows']) - len(validation['missing_bom']):,}개")
    c4.metric("BOM 없음", f"{len(validation['missing_bom']):,}개")

    section("생산 대상 확인", "아래 표의 모든 행이 정상이어야 전체 생산을 실행할 수 있습니다.")
    preview = _preview_frame(pd, validation["rows"])
    st.dataframe(
        preview,
        use_container_width=True,
        hide_index=True,
        height=min(650, max(220, 38 * (len(preview) + 1))),
    )

    if duplicate:
        st.error(
            f"이 파일은 이미 생산 처리되었습니다. 생산일 {duplicate['production_date']} · "
            f"처리시각 {duplicate['created_at']} · 동일 파일 재생산은 차단됩니다."
        )
        return

    if validation["missing_bom"]:
        st.error(
            f"BOM이 없는 상품이 {len(validation['missing_bom']):,}개 있습니다. "
            "요청하신 규칙에 따라 정상 상품을 포함한 전체 생산을 중단합니다."
        )
        missing = pd.DataFrame([
            {
                "상품명": _text(r.get("source_name")),
                "옵션명": _text(r.get("option_name")),
                "옵션ID": _text(r.get("option_id")),
                "ERP 상품명": _text(r.get("erp_name")),
            }
            for r in validation["missing_bom"]
        ])
        st.dataframe(missing, use_container_width=True, hide_index=True)
        st.warning("BOM을 등록한 뒤 생산자료를 다시 업로드해 주세요. 현재 파일에서는 어떤 상품도 생산되지 않았습니다.")
        return

    if not validation["ok"]:
        st.error(
            f"생산자료에 오류가 있는 상품이 {len(validation['blocked']):,}개 있습니다. "
            "전체 생산을 중단합니다. 상태 칸의 오류를 수정한 뒤 파일을 다시 업로드해 주세요."
        )
        return

    st.success("모든 생산대상 상품의 옵션ID와 BOM 확인이 완료되었습니다. 아직 생산은 실행되지 않았습니다.")
    confirm = st.checkbox(
        f"위 {len(validation['rows']):,}개 상품 / 총 {validation['total_qty']:,}개를 {production_date:%Y-%m-%d} 생산 처리합니다.",
        key=f"production_batch_v095_confirm_{parsed['file_hash'][:12]}",
    )
    if st.button(
        "전체 생산 실행",
        type="primary",
        disabled=not confirm,
        key=f"production_batch_v095_execute_{parsed['file_hash'][:12]}",
    ):
        try:
            result = execute_batch(core, parsed, getattr(uploaded, "name", "생산자료.xlsx"), production_date, db_path)
            st.success(
                f"생산 완료: {result['rows']:,}개 상품 / 총 {result['total_qty']:,}개. "
                "BOM 구성품은 자체창고에서 차감되고 완제품은 쿠팡RG에 입고되었습니다."
            )
            st.rerun()
        except Exception as exc:
            st.error(f"생산 실패: {exc}")


def patch_source(source: str) -> str:
    menu_label = '        "🏭  생산자료",\n'
    if menu_label not in source:
        anchor = '        "📦  재고관리",\n'
        if anchor not in source:
            raise RuntimeError("생산자료 메뉴를 추가할 위치를 찾지 못했습니다.")
        source = source.replace(anchor, menu_label + anchor, 1)

    handler = '''# ------------------------------
# Batch production Excel
# ------------------------------
elif page == "🏭  생산자료":
    production_batch_v095.render_production_batch_page(
        st=st, pd=pd, date=date, core=core, page_header=page_header, section=section,
        kpi=kpi, money=money, fmt_date=fmt_date, latest_updated_text=latest_updated_text,
    )


'''
    if 'elif page == "🏭  생산자료":' not in source:
        anchor = '# ------------------------------\n# Inventory\n# ------------------------------\nelif page == "📦  재고관리":\n'
        if anchor not in source:
            raise RuntimeError("생산자료 화면을 추가할 위치를 찾지 못했습니다.")
        source = source.replace(anchor, handler + anchor, 1)
    return source
