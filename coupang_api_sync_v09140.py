"""Manual Coupang Open API synchronization for RG Manager v0.9.142.

The module deliberately performs no network work at import/startup.  Every API
request is initiated by an explicit Streamlit button click.

Supported official endpoints:
- Rocket Growth order list
- Rocket warehouse inventory summaries
- Revenue/sales details
- Settlement/payment histories

Raw API facts are preserved in dedicated SQLite tables.  Rows are linked to the
existing product master by immutable Coupang vendorItemId (ERP option_id).  The
inventory action separates normal options from explicitly confirmed returned-item
aliases.  The intersection of Rocket Growth inbound Excel option ids and active
ERP product codes is kept as the normal-option registry.  Normal stock is reconciled to ``쿠팡RG``
while all return-option quantities belonging to one original product are summed
and reconciled once to ``반품창고``.
"""
from __future__ import annotations

import base64
import calendar
import ctypes
from ctypes import wintypes
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import hashlib
import hmac
import json
import os
from pathlib import Path
import functools
import re
import sqlite3
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen


PAGE_LABEL = "🔗  쿠팡 API 연동"
API_HOST = "https://api-gateway.coupang.com"
CONFIG_FILE = "coupang_api_credentials.dat"
_MARKER = "# _rg_coupang_api_sync_v09140"


class CoupangAPIError(RuntimeError):
    """A user-facing, secret-safe Coupang API failure."""


@dataclass(frozen=True)
class Credentials:
    vendor_id: str
    access_key: str
    secret_key: str

    def validate(self) -> None:
        if not self.vendor_id.strip().startswith("A"):
            raise ValueError("판매자 ID는 A로 시작하는 값을 입력해 주세요.")
        if not self.access_key.strip():
            raise ValueError("Access Key를 입력해 주세요.")
        if not self.secret_key.strip():
            raise ValueError("Secret Key를 입력해 주세요.")


def _num(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _integer(value: Any, default: int = 0) -> int:
    try:
        return int(round(float(value)))
    except Exception:
        return default


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _oid(value: Any) -> str:
    text = _text(value)
    if text.upper().startswith("CP-"):
        text = text[3:]
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _local_now(core: Any | None = None) -> str:
    if core is not None:
        try:
            return str(core.now_iso())
        except Exception:
            pass
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _to_date(value: date | str) -> date:
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _date_chunks(start: date | str, end: date | str, max_days: int):
    left, right = _to_date(start), _to_date(end)
    if left > right:
        raise ValueError("조회 시작일은 종료일보다 늦을 수 없습니다.")
    cursor = left
    while cursor <= right:
        chunk_end = min(right, cursor + timedelta(days=max_days - 1))
        yield cursor, chunk_end
        cursor = chunk_end + timedelta(days=1)


def _extract_rows(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if not isinstance(payload, dict):
        return []
    data = payload.get("data")
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for key in ("content", "items", "orders", "results"):
            value = data.get(key)
            if isinstance(value, list):
                return [x for x in value if isinstance(x, dict)]
    for key in ("content", "items", "orders", "results"):
        value = payload.get(key)
        if isinstance(value, list):
            return [x for x in value if isinstance(x, dict)]
    return []


def _next_token(payload: Any) -> str:
    if not isinstance(payload, dict):
        return ""
    for container in (payload, payload.get("data")):
        if not isinstance(container, dict):
            continue
        for key in ("nextToken", "next_token", "token"):
            value = _text(container.get(key))
            if value:
                return value
    return ""


class CoupangClient:
    def __init__(
        self,
        credentials: Credentials,
        opener: Callable[..., Any] = urlopen,
        now: Callable[[], datetime] | None = None,
        timeout: int = 30,
    ):
        credentials.validate()
        self.credentials = credentials
        self.opener = opener
        self.now = now or (lambda: datetime.now(timezone.utc))
        self.timeout = int(timeout)

    def _signed_date(self) -> str:
        current = self.now()
        if current.tzinfo is None:
            current = current.replace(tzinfo=timezone.utc)
        return current.astimezone(timezone.utc).strftime("%y%m%dT%H%M%SZ")

    def authorization(self, method: str, path: str, query: str, signed_date: str | None = None) -> str:
        stamp = signed_date or self._signed_date()
        message = stamp + method.upper() + path + query
        signature = hmac.new(
            self.credentials.secret_key.encode("utf-8"),
            message.encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        return (
            "CEA algorithm=HmacSHA256, "
            f"access-key={self.credentials.access_key}, "
            f"signed-date={stamp}, signature={signature}"
        )

    def request(self, path: str, params: Iterable[tuple[str, Any]] | dict[str, Any] | None = None) -> Any:
        items = list(params.items()) if isinstance(params, dict) else list(params or [])
        clean = [(str(k), str(v)) for k, v in items if v is not None]
        query = urlencode(clean)
        stamp = self._signed_date()
        uri = path + (("?" + query) if query else "")
        req = Request(
            API_HOST + uri,
            method="GET",
            headers={
                "Authorization": self.authorization("GET", path, query, stamp),
                "X-Requested-By": self.credentials.vendor_id,
                "Content-Type": "application/json;charset=UTF-8",
                "User-Agent": "RG-Manager/0.9.140",
            },
        )
        try:
            response = self.opener(req, timeout=self.timeout)
            raw = response.read()
            return json.loads(raw.decode("utf-8")) if raw else {}
        except HTTPError as exc:
            try:
                body = exc.read().decode("utf-8", errors="replace")
                parsed = json.loads(body)
                detail = _text(parsed.get("message") or parsed.get("error"))
            except Exception:
                detail = ""
            suffix = f": {detail}" if detail else ""
            raise CoupangAPIError(f"쿠팡 API 요청 실패 (HTTP {exc.code}){suffix}") from None
        except URLError as exc:
            raise CoupangAPIError(f"쿠팡 API에 연결할 수 없습니다: {_text(exc.reason)}") from None
        except TimeoutError:
            raise CoupangAPIError("쿠팡 API 응답 시간이 초과되었습니다.") from None
        except json.JSONDecodeError:
            raise CoupangAPIError("쿠팡 API 응답을 읽을 수 없습니다.") from None

    def _paged(self, path: str, params: list[tuple[str, Any]], token_name: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        token = ""
        seen: set[str] = set()
        for _page in range(500):
            request_params = [(k, v) for k, v in params if k != token_name]
            if token or token_name == "token":
                request_params.append((token_name, token))
            payload = self.request(path, request_params)
            rows.extend(_extract_rows(payload))
            nxt = _next_token(payload)
            if not nxt:
                return rows
            if nxt in seen:
                raise CoupangAPIError("쿠팡 API가 동일한 다음 페이지 값을 반복했습니다.")
            seen.add(nxt)
            token = nxt
        raise CoupangAPIError("쿠팡 API 페이지 수가 안전 한도를 초과했습니다.")

    def orders(self, start: date | str, end: date | str) -> list[dict[str, Any]]:
        path = (
            "/v2/providers/rg_open_api/apis/api/v1/vendors/"
            f"{self.credentials.vendor_id}/rg/orders"
        )
        rows: list[dict[str, Any]] = []
        for a, b in _date_chunks(start, end, 30):
            rows.extend(self._paged(path, [
                ("paidDateFrom", a.strftime("%Y%m%d")),
                ("paidDateTo", b.strftime("%Y%m%d")),
            ], "nextToken"))
        return rows

    def inventory(self) -> list[dict[str, Any]]:
        path = (
            "/v2/providers/rg_open_api/apis/api/v1/vendors/"
            f"{self.credentials.vendor_id}/rg/inventory/summaries"
        )
        return self._paged(path, [], "nextToken")

    def revenue(self, start: date | str, end: date | str) -> list[dict[str, Any]]:
        path = "/v2/providers/openapi/apis/api/v1/revenue-history"
        rows: list[dict[str, Any]] = []
        for a, b in _date_chunks(start, end, 31):
            rows.extend(self._paged(path, [
                ("vendorId", self.credentials.vendor_id),
                ("recognitionDateFrom", a.isoformat()),
                ("recognitionDateTo", b.isoformat()),
                ("maxPerPage", 50),
            ], "token"))
        return rows

    def settlements(self, month: str) -> list[dict[str, Any]]:
        path = "/v2/providers/marketplace_openapi/apis/api/v1/settlement-histories"
        payload = self.request(path, [("revenueRecognitionYearMonth", str(month))])
        return _extract_rows(payload) if isinstance(payload, dict) else _extract_rows(payload)


class _DataBlob(ctypes.Structure):
    _fields_ = [("cbData", wintypes.DWORD), ("pbData", ctypes.POINTER(ctypes.c_byte))]


def _blob(data: bytes):
    buffer = ctypes.create_string_buffer(data)
    return _DataBlob(len(data), ctypes.cast(buffer, ctypes.POINTER(ctypes.c_byte))), buffer


def _dpapi_encrypt(data: bytes) -> bytes:
    if os.name != "nt":
        raise RuntimeError("API 키 암호화 저장은 Windows 실행 환경에서만 지원합니다.")
    in_blob, in_buffer = _blob(data)
    entropy_blob, entropy_buffer = _blob(b"RG-Manager-Coupang-API-v09140")
    out_blob = _DataBlob()
    protect = ctypes.windll.crypt32.CryptProtectData
    protect.argtypes = [
        ctypes.POINTER(_DataBlob), wintypes.LPCWSTR, ctypes.POINTER(_DataBlob),
        ctypes.c_void_p, ctypes.c_void_p, wintypes.DWORD, ctypes.POINTER(_DataBlob),
    ]
    protect.restype = wintypes.BOOL
    ok = protect(
        ctypes.byref(in_blob),
        "RG Manager Coupang API",
        ctypes.byref(entropy_blob),
        None,
        None,
        0,
        ctypes.byref(out_blob),
    )
    _ = (in_buffer, entropy_buffer)
    if not ok:
        raise ctypes.WinError()
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        local_free = ctypes.windll.kernel32.LocalFree
        local_free.argtypes = [ctypes.c_void_p]
        local_free.restype = ctypes.c_void_p
        local_free(ctypes.cast(out_blob.pbData, ctypes.c_void_p))


def _dpapi_decrypt(data: bytes) -> bytes:
    if os.name != "nt":
        raise RuntimeError("API 키 암호화 해제는 Windows 실행 환경에서만 지원합니다.")
    in_blob, in_buffer = _blob(data)
    entropy_blob, entropy_buffer = _blob(b"RG-Manager-Coupang-API-v09140")
    out_blob = _DataBlob()
    unprotect = ctypes.windll.crypt32.CryptUnprotectData
    unprotect.argtypes = [
        ctypes.POINTER(_DataBlob), ctypes.POINTER(wintypes.LPWSTR),
        ctypes.POINTER(_DataBlob), ctypes.c_void_p, ctypes.c_void_p,
        wintypes.DWORD, ctypes.POINTER(_DataBlob),
    ]
    unprotect.restype = wintypes.BOOL
    ok = unprotect(
        ctypes.byref(in_blob), None, ctypes.byref(entropy_blob), None, None, 0, ctypes.byref(out_blob)
    )
    _ = (in_buffer, entropy_buffer)
    if not ok:
        raise ctypes.WinError()
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        local_free = ctypes.windll.kernel32.LocalFree
        local_free.argtypes = [ctypes.c_void_p]
        local_free.restype = ctypes.c_void_p
        local_free(ctypes.cast(out_blob.pbData, ctypes.c_void_p))


def _config_path(core: Any) -> Path:
    db = Path(str(core.DEFAULT_DB)).resolve()
    return db.parent / CONFIG_FILE


def save_credentials(core: Any, credentials: Credentials) -> None:
    credentials.validate()
    payload = _json({
        "vendor_id": credentials.vendor_id.strip(),
        "access_key": credentials.access_key.strip(),
        "secret_key": credentials.secret_key.strip(),
    }).encode("utf-8")
    encrypted = base64.b64encode(_dpapi_encrypt(payload))
    path = _config_path(core)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_bytes(encrypted)
    tmp.replace(path)


def load_credentials(core: Any) -> Credentials | None:
    env = Credentials(
        _text(os.getenv("COUPANG_VENDOR_ID")),
        _text(os.getenv("COUPANG_ACCESS_KEY")),
        _text(os.getenv("COUPANG_SECRET_KEY")),
    )
    if env.vendor_id and env.access_key and env.secret_key:
        return env
    path = _config_path(core)
    if not path.exists():
        return None
    try:
        payload = json.loads(_dpapi_decrypt(base64.b64decode(path.read_bytes())).decode("utf-8"))
        result = Credentials(
            _text(payload.get("vendor_id")),
            _text(payload.get("access_key")),
            _text(payload.get("secret_key")),
        )
        result.validate()
        return result
    except Exception as exc:
        raise RuntimeError("저장된 쿠팡 API 키를 읽을 수 없습니다. 연결정보를 다시 저장해 주세요.") from exc


def delete_credentials(core: Any) -> None:
    path = _config_path(core)
    if path.exists():
        path.unlink()


def _exists(con: sqlite3.Connection, table: str) -> bool:
    return con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
    ).fetchone() is not None


def _columns(con: sqlite3.Connection, table: str) -> set[str]:
    if not _exists(con, table):
        return set()
    return {str(r["name"]) for r in con.execute(f'PRAGMA table_info("{table}")').fetchall()}


def ensure_schema(core: Any, db_path: Any | None = None) -> None:
    db = db_path or core.DEFAULT_DB
    core.init_db(db)
    with core._conn(db) as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS coupang_api_sync_runs(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sync_type TEXT NOT NULL,
                period_start TEXT,
                period_end TEXT,
                status TEXT NOT NULL,
                rows_received INTEGER NOT NULL DEFAULT 0,
                rows_matched INTEGER NOT NULL DEFAULT 0,
                message TEXT,
                started_at TEXT NOT NULL,
                completed_at TEXT
            );
            CREATE INDEX IF NOT EXISTS ix_coupang_api_sync_runs_started
              ON coupang_api_sync_runs(started_at DESC);

            CREATE TABLE IF NOT EXISTS coupang_rg_order_items(
                order_id TEXT NOT NULL,
                item_index INTEGER NOT NULL,
                paid_at TEXT,
                paid_date TEXT,
                vendor_item_id TEXT NOT NULL,
                product_id INTEGER,
                product_name TEXT,
                sales_quantity REAL NOT NULL DEFAULT 0,
                unit_sales_price REAL NOT NULL DEFAULT 0,
                currency TEXT,
                raw_json TEXT NOT NULL,
                synced_at TEXT NOT NULL,
                PRIMARY KEY(order_id,item_index)
            );
            CREATE INDEX IF NOT EXISTS ix_coupang_rg_order_items_paid
              ON coupang_rg_order_items(paid_date,vendor_item_id);

            CREATE TABLE IF NOT EXISTS coupang_rg_inventory(
                vendor_item_id TEXT PRIMARY KEY,
                product_id INTEGER,
                external_sku_id TEXT,
                orderable_qty REAL NOT NULL DEFAULT 0,
                sales_30d REAL NOT NULL DEFAULT 0,
                raw_json TEXT NOT NULL,
                synced_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS coupang_rg_inventory_snapshots(
                run_id INTEGER NOT NULL,
                vendor_item_id TEXT NOT NULL,
                product_id INTEGER,
                orderable_qty REAL NOT NULL DEFAULT 0,
                sales_30d REAL NOT NULL DEFAULT 0,
                captured_at TEXT NOT NULL,
                PRIMARY KEY(run_id,vendor_item_id)
            );

            CREATE TABLE IF NOT EXISTS coupang_revenue_items(
                order_id TEXT NOT NULL,
                sale_type TEXT NOT NULL,
                recognition_date TEXT NOT NULL,
                transaction_index INTEGER NOT NULL,
                item_index INTEGER NOT NULL,
                sale_date TEXT,
                settlement_date TEXT,
                vendor_item_id TEXT NOT NULL,
                product_id INTEGER,
                product_name TEXT,
                vendor_item_name TEXT,
                quantity REAL NOT NULL DEFAULT 0,
                sale_price REAL NOT NULL DEFAULT 0,
                coupang_discount_coupon REAL NOT NULL DEFAULT 0,
                seller_discount_coupon REAL NOT NULL DEFAULT 0,
                downloadable_coupon REAL NOT NULL DEFAULT 0,
                sale_amount REAL NOT NULL DEFAULT 0,
                service_fee REAL NOT NULL DEFAULT 0,
                service_fee_vat REAL NOT NULL DEFAULT 0,
                settlement_amount REAL NOT NULL DEFAULT 0,
                raw_json TEXT NOT NULL,
                synced_at TEXT NOT NULL,
                PRIMARY KEY(order_id,sale_type,recognition_date,transaction_index,item_index)
            );
            CREATE INDEX IF NOT EXISTS ix_coupang_revenue_items_recognition
              ON coupang_revenue_items(recognition_date,vendor_item_id);

            CREATE TABLE IF NOT EXISTS coupang_settlement_histories(
                revenue_month TEXT NOT NULL,
                settlement_type TEXT NOT NULL,
                settlement_date TEXT NOT NULL,
                item_index INTEGER NOT NULL,
                recognition_date_from TEXT,
                recognition_date_to TEXT,
                total_sale REAL NOT NULL DEFAULT 0,
                service_fee REAL NOT NULL DEFAULT 0,
                settlement_target_amount REAL NOT NULL DEFAULT 0,
                settlement_amount REAL NOT NULL DEFAULT 0,
                last_amount REAL NOT NULL DEFAULT 0,
                pending_released_amount REAL NOT NULL DEFAULT 0,
                deduction_amount REAL NOT NULL DEFAULT 0,
                final_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                raw_json TEXT NOT NULL,
                synced_at TEXT NOT NULL,
                PRIMARY KEY(revenue_month,settlement_type,settlement_date,item_index)
            );

            CREATE TABLE IF NOT EXISTS coupang_api_option_catalog(
                vendor_item_id TEXT PRIMARY KEY,
                product_name TEXT,
                vendor_item_name TEXT,
                external_sku_id TEXT,
                source TEXT,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS coupang_normal_option_registry(
                vendor_item_id TEXT PRIMARY KEY,
                product_name TEXT,
                option_name TEXT,
                exposure_product_id TEXT,
                seller_product_id TEXT,
                source_file TEXT,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS return_discount_aliases(
                discount_option_id TEXT PRIMARY KEY,
                parent_product_id INTEGER NOT NULL,
                discount_name TEXT,
                match_method TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )
        # v0.9.141 keeps the API option row distinct even when multiple returned
        # options roll up to one original ERP product.
        if "stock_type" not in _columns(con, "coupang_rg_inventory"):
            con.execute(
                "ALTER TABLE coupang_rg_inventory "
                "ADD COLUMN stock_type TEXT NOT NULL DEFAULT 'unmatched'"
            )
        if "stock_type" not in _columns(con, "coupang_rg_inventory_snapshots"):
            con.execute(
                "ALTER TABLE coupang_rg_inventory_snapshots "
                "ADD COLUMN stock_type TEXT NOT NULL DEFAULT 'unmatched'"
            )


def parse_rg_inbound_options(source: Any) -> list[dict[str, str]]:
    """Read the official RG inbound workbook and return its normal option ids."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"Excel 읽기 모듈(openpyxl)을 불러오지 못했습니다: {exc}") from None

    if isinstance(source, (str, os.PathLike, Path)):
        workbook_source: Any = str(source)
    else:
        raw = source.getvalue() if hasattr(source, "getvalue") else source.read()
        workbook_source = BytesIO(raw)
    try:
        book = load_workbook(workbook_source, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"입고 Excel 파일을 읽지 못했습니다: {exc}") from None

    aliases = {
        "product_name": {"등록상품명", "상품명"},
        "option_name": {"옵션명"},
        "exposure_product_id": {"노출상품 ID", "노출상품ID"},
        "seller_product_id": {"등록상품 ID", "등록상품ID"},
        "vendor_item_id": {"옵션 ID", "옵션ID"},
        "sale_method": {"판매 방식", "판매방식"},
    }
    selected = None
    header_row = 0
    header_cols: dict[str, int] = {}
    for sheet in book.worksheets:
        for row_no, values in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            normalized = {_text(value).replace("\n", " "): idx for idx, value in enumerate(values)}
            found: dict[str, int] = {}
            for key, names in aliases.items():
                for name in names:
                    if name in normalized:
                        found[key] = normalized[name]
                        break
            if "vendor_item_id" in found:
                selected, header_row, header_cols = sheet, row_no, found
                break
        if selected is not None:
            break
    if selected is None:
        raise ValueError("G열의 '옵션 ID' 머리글을 찾지 못했습니다. 쿠팡 로켓그로스 입고 Excel인지 확인해 주세요.")

    rows_by_id: dict[str, dict[str, str]] = {}
    for values in selected.iter_rows(min_row=header_row + 1, values_only=True):
        def value(key: str) -> Any:
            idx = header_cols.get(key)
            return values[idx] if idx is not None and idx < len(values) else ""

        oid = _oid(value("vendor_item_id"))
        if not oid:
            continue
        sale_method = _text(value("sale_method"))
        if sale_method and "로켓그로스" not in sale_method:
            continue
        rows_by_id[oid] = {
            "vendor_item_id": oid,
            "product_name": _text(value("product_name")),
            "option_name": _text(value("option_name")),
            "exposure_product_id": _oid(value("exposure_product_id")),
            "seller_product_id": _oid(value("seller_product_id")),
        }
    if not rows_by_id:
        raise ValueError("입고 Excel에서 로켓그로스 옵션 ID를 찾지 못했습니다.")
    return list(rows_by_id.values())


def _normal_option_ids(con: sqlite3.Connection) -> set[str]:
    if not _exists(con, "coupang_normal_option_registry"):
        return set()
    return {
        _oid(row["vendor_item_id"])
        for row in con.execute("SELECT vendor_item_id FROM coupang_normal_option_registry")
        if _oid(row["vendor_item_id"])
    }


def _erp_direct_product_map(
    con: sqlite3.Connection, active_only: bool = False
) -> dict[str, int]:
    """Map option/item codes physically present in the ERP product master."""
    if not _exists(con, "products"):
        return {}
    cols = _columns(con, "products")
    if not {"id", "option_id"}.issubset(cols):
        return {}
    select_code = ",item_code" if "item_code" in cols else ",'' AS item_code"
    where = " WHERE COALESCE(active,1)=1" if active_only and "active" in cols else ""
    out: dict[str, int] = {}
    for row in con.execute("SELECT id,option_id" + select_code + " FROM products" + where):
        for key in (_oid(row["option_id"]), _oid(row["item_code"])):
            if key and key.isdigit():
                out.setdefault(key, int(row["id"]))
    return out


def register_normal_options(
    core: Any,
    rows: Iterable[dict[str, Any]],
    source_file: str = "",
    db_path: Any | None = None,
) -> dict[str, int]:
    """Replace the normal registry with inbound ids matching active ERP products."""
    db = db_path or core.DEFAULT_DB
    ensure_schema(core, db)
    cleaned: dict[str, dict[str, str]] = {}
    for row in rows:
        oid = _oid(row.get("vendor_item_id"))
        if not oid:
            continue
        cleaned[oid] = {
            "product_name": _text(row.get("product_name")),
            "option_name": _text(row.get("option_name")),
            "exposure_product_id": _oid(row.get("exposure_product_id")),
            "seller_product_id": _oid(row.get("seller_product_id")),
        }
    if not cleaned:
        raise ValueError("등록할 정상상품 옵션 ID가 없습니다.")

    with core._conn(db) as con:
        now = _local_now(core)
        before = _normal_option_ids(con)
        direct_mapping = _erp_direct_product_map(con, active_only=True)
        verified = {oid: row for oid, row in cleaned.items() if oid in direct_mapping}
        removed = len(before - set(verified))
        if verified:
            keep_placeholders = ",".join("?" for _ in verified)
            con.execute(
                f"DELETE FROM coupang_normal_option_registry WHERE vendor_item_id NOT IN ({keep_placeholders})",
                tuple(verified),
            )
        else:
            con.execute("DELETE FROM coupang_normal_option_registry")
        alias_conflicts = 0
        if verified:
            con.executemany(
                """INSERT INTO coupang_normal_option_registry
                   (vendor_item_id,product_name,option_name,exposure_product_id,
                    seller_product_id,source_file,first_seen_at,last_seen_at)
                   VALUES(?,?,?,?,?,?,?,?)
                   ON CONFLICT(vendor_item_id) DO UPDATE SET
                     product_name=CASE WHEN excluded.product_name<>'' THEN excluded.product_name
                                       ELSE coupang_normal_option_registry.product_name END,
                     option_name=CASE WHEN excluded.option_name<>'' THEN excluded.option_name
                                      ELSE coupang_normal_option_registry.option_name END,
                     exposure_product_id=CASE WHEN excluded.exposure_product_id<>'' THEN excluded.exposure_product_id
                                              ELSE coupang_normal_option_registry.exposure_product_id END,
                     seller_product_id=CASE WHEN excluded.seller_product_id<>'' THEN excluded.seller_product_id
                                            ELSE coupang_normal_option_registry.seller_product_id END,
                     source_file=excluded.source_file,
                     last_seen_at=excluded.last_seen_at""",
                [(
                    oid, row["product_name"], row["option_name"], row["exposure_product_id"],
                    row["seller_product_id"], _text(source_file), now, now,
                ) for oid, row in verified.items()],
            )
            placeholders = ",".join("?" for _ in verified)
            alias_conflicts = int(con.execute(
                f"SELECT COUNT(*) n FROM return_discount_aliases WHERE discount_option_id IN ({placeholders})",
                tuple(verified),
            ).fetchone()["n"])
            con.execute(
                f"DELETE FROM return_discount_aliases WHERE discount_option_id IN ({placeholders})",
                tuple(verified),
            )
        total = int(con.execute(
            "SELECT COUNT(*) n FROM coupang_normal_option_registry"
        ).fetchone()["n"])
    return {
        "rows": len(cleaned),
        "registered": len(verified),
        "new": len(set(verified) - before),
        "removed": removed,
        "matched": len(verified),
        "unmatched": len(cleaned) - len(verified),
        "alias_conflicts_removed": alias_conflicts,
        "total": total,
    }


def _product_map(con: sqlite3.Connection) -> dict[str, int]:
    out = _erp_direct_product_map(con)
    if _exists(con, "return_discount_aliases"):
        official_normal = _normal_option_ids(con)
        cols = _columns(con, "return_discount_aliases")
        if {"discount_option_id", "parent_product_id"}.issubset(cols):
            for row in con.execute(
                "SELECT discount_option_id,parent_product_id FROM return_discount_aliases"
            ):
                key = _oid(row["discount_option_id"])
                if key and key not in official_normal:
                    # An explicit return alias must override a legacy auto-created
                    # child product carrying the same option id.
                    out[key] = int(row["parent_product_id"])
    return out


def _normal_product_map(
    con: sqlite3.Connection, active_only: bool = False
) -> dict[str, int]:
    """Return only managed normal-product option ids, never return aliases."""
    if not _exists(con, "products"):
        return {}
    cols = _columns(con, "products")
    if not {"id", "option_id"}.issubset(cols):
        return {}
    official_normal = _normal_option_ids(con)
    aliases = {
        _oid(r["discount_option_id"])
        for r in con.execute("SELECT discount_option_id FROM return_discount_aliases")
    } - official_normal
    select_code = ",item_code" if "item_code" in cols else ",'' AS item_code"
    where = " WHERE COALESCE(active,1)=1" if active_only and "active" in cols else ""
    out: dict[str, int] = {}
    for row in con.execute("SELECT id,option_id" + select_code + " FROM products" + where):
        option_id = _oid(row["option_id"])
        if option_id and option_id not in aliases:
            out.setdefault(option_id, int(row["id"]))
        item_code = _oid(row["item_code"])
        if item_code and item_code.isdigit() and item_code not in aliases:
            out.setdefault(item_code, int(row["id"]))
    return out


def _return_product_map(con: sqlite3.Connection) -> dict[str, int]:
    if not _exists(con, "return_discount_aliases"):
        return {}
    return {
        _oid(r["discount_option_id"]): int(r["parent_product_id"])
        for r in con.execute(
            "SELECT discount_option_id,parent_product_id FROM return_discount_aliases"
        )
        if _oid(r["discount_option_id"])
    }


def _name_key(value: Any) -> str:
    """Match the existing return-sale name rule without guessing variants."""
    text = _text(value).lower()
    text = re.sub(
        r"[,/\s]+\d+\s*(?:개입|개|p|pcs?|세트|set)\s*$", "", text, flags=re.I
    )
    text = re.sub(r"\s+", " ", text).strip(" ,-/")
    return re.sub(r"[\s,·_/\-]+", "", text)


def _catalog_upsert(
    con: sqlite3.Connection,
    observations: Iterable[dict[str, Any]],
    now: str,
) -> None:
    rows = []
    for row in observations:
        oid = _oid(row.get("vendor_item_id"))
        if not oid:
            continue
        rows.append((
            oid,
            _text(row.get("product_name")),
            _text(row.get("vendor_item_name")),
            _text(row.get("external_sku_id")),
            _text(row.get("source")),
            now,
        ))
    con.executemany(
        """INSERT INTO coupang_api_option_catalog
           (vendor_item_id,product_name,vendor_item_name,external_sku_id,source,updated_at)
           VALUES(?,?,?,?,?,?)
           ON CONFLICT(vendor_item_id) DO UPDATE SET
             product_name=CASE WHEN excluded.product_name<>'' THEN excluded.product_name
                               ELSE coupang_api_option_catalog.product_name END,
             vendor_item_name=CASE WHEN excluded.vendor_item_name<>'' THEN excluded.vendor_item_name
                                   ELSE coupang_api_option_catalog.vendor_item_name END,
             external_sku_id=CASE WHEN excluded.external_sku_id<>'' THEN excluded.external_sku_id
                                  ELSE coupang_api_option_catalog.external_sku_id END,
             source=CASE WHEN excluded.source<>'' THEN excluded.source
                         ELSE coupang_api_option_catalog.source END,
             updated_at=excluded.updated_at""",
        rows,
    )


def _normal_name_candidates(con: sqlite3.Connection) -> dict[str, set[int]]:
    if not _exists(con, "products"):
        return {}
    cols = _columns(con, "products")
    if not {"id", "name", "option_id"}.issubset(cols):
        return {}
    fields = ["id", "name", "option_id"]
    fields.append("item_code" if "item_code" in cols else "'' AS item_code")
    fields.append("unit_cost" if "unit_cost" in cols else "0 AS unit_cost")
    active_where = " WHERE COALESCE(active,1)=1" if "active" in cols else ""
    official_ids = _normal_option_ids(con)
    if not official_ids:
        return {}
    alias_ids = set(_return_product_map(con)) - official_ids
    out: dict[str, set[int]] = {}
    for row in con.execute("SELECT " + ",".join(fields) + " FROM products" + active_where):
        oid = _oid(row["option_id"])
        code = _oid(row["item_code"])
        if not oid or oid in alias_ids or not ({oid, code} & official_ids):
            continue
        # Do not let a legacy zero-cost CP-{optionId} child become its own
        # original-product candidate.
        raw_code = _text(row["item_code"])
        if raw_code.upper() in {oid.upper(), ("CP-" + oid).upper()} and abs(_num(row["unit_cost"])) <= 1e-9:
            continue
        key = _name_key(row["name"])
        if key:
            out.setdefault(key, set()).add(int(row["id"]))
    return out


def _auto_link_verified_return_aliases(
    con: sqlite3.Connection,
    option_ids: Iterable[str] | None,
    now: str,
) -> int:
    """Auto-link only against a unique parent proven normal by inbound Excel."""
    official_ids = _normal_option_ids(con)
    if not official_ids:
        return 0
    normal_ids = set(_normal_product_map(con)) | official_ids
    return_ids = set(_return_product_map(con)) - official_ids
    candidates = _normal_name_candidates(con)
    wanted = {_oid(x) for x in option_ids or [] if _oid(x)}
    linked = 0
    for oid in sorted(wanted):
        if oid in normal_ids or oid in return_ids:
            continue
        catalog = con.execute(
            """SELECT product_name,vendor_item_name
               FROM coupang_api_option_catalog WHERE vendor_item_id=?""",
            (oid,),
        ).fetchone()
        if not catalog:
            continue
        names = [_text(catalog["product_name"]), _text(catalog["vendor_item_name"])]
        matched: set[int] = set()
        for name in names:
            key = _name_key(name)
            if key:
                matched.update(candidates.get(key, set()))
        if len(matched) != 1:
            continue
        parent_id = next(iter(matched))
        display_name = next((name for name in names if name), "")
        con.execute(
            """INSERT INTO return_discount_aliases
               (discount_option_id,parent_product_id,discount_name,match_method,created_at,updated_at)
               VALUES(?,?,?,?,?,?)
               ON CONFLICT(discount_option_id) DO NOTHING""",
            (oid, parent_id, display_name, "api_verified_normal_name", now, now),
        )
        linked += int(con.execute("SELECT changes() n").fetchone()["n"] or 0)
    return linked


def save_return_mapping(
    core: Any,
    discount_option_id: str,
    parent_product_id: int,
    discount_name: str = "",
    db_path: Any | None = None,
) -> None:
    """Save an explicit one-time return option -> original product mapping."""
    db = db_path or core.DEFAULT_DB
    ensure_schema(core, db)
    oid = _oid(discount_option_id)
    if not oid:
        raise ValueError("반품 옵션ID를 선택해 주세요.")
    with core._conn(db) as con:
        if oid in _normal_option_ids(con):
            raise ValueError(
                "입고 Excel에 등록된 정상상품 옵션 ID입니다. 반품상품으로 연결할 수 없습니다."
            )
        parent = con.execute(
            "SELECT id FROM products WHERE id=?", (int(parent_product_id),)
        ).fetchone()
        if not parent:
            raise ValueError("연결할 원상품을 찾지 못했습니다.")
        now = _local_now(core)
        con.execute(
            """INSERT INTO return_discount_aliases
               (discount_option_id,parent_product_id,discount_name,match_method,created_at,updated_at)
               VALUES(?,?,?,?,?,?)
               ON CONFLICT(discount_option_id) DO UPDATE SET
                 parent_product_id=excluded.parent_product_id,
                 discount_name=excluded.discount_name,
                 match_method=excluded.match_method,
                 updated_at=excluded.updated_at""",
            (oid, int(parent_product_id), _text(discount_name), "api_manual", now, now),
        )


def _run_start(core: Any, db: Any, kind: str, start: str | None, end: str | None) -> int:
    ensure_schema(core, db)
    with core._conn(db) as con:
        cur = con.execute(
            """INSERT INTO coupang_api_sync_runs
               (sync_type,period_start,period_end,status,started_at)
               VALUES(?,?,?,?,?)""",
            (kind, start, end, "running", _local_now(core)),
        )
        return int(cur.lastrowid)


def _run_finish(
    core: Any,
    db: Any,
    run_id: int,
    status: str,
    received: int = 0,
    matched: int = 0,
    message: str = "",
) -> None:
    with core._conn(db) as con:
        con.execute(
            """UPDATE coupang_api_sync_runs
               SET status=?,rows_received=?,rows_matched=?,message=?,completed_at=?
               WHERE id=?""",
            (status, int(received), int(matched), _text(message), _local_now(core), int(run_id)),
        )


def _paid_parts(value: Any) -> tuple[str, str]:
    text = _text(value)
    if text.isdigit():
        try:
            dt = datetime.fromtimestamp(int(text) / 1000, tz=timezone.utc)
            return dt.isoformat(), dt.date().isoformat()
        except Exception:
            pass
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return dt.isoformat(), dt.date().isoformat()
    except Exception:
        return text, text[:10] if len(text) >= 10 else ""


def sync_orders(core: Any, client: CoupangClient, start: date | str, end: date | str, db_path=None):
    db = db_path or core.DEFAULT_DB
    a, b = _to_date(start), _to_date(end)
    run_id = _run_start(core, db, "orders", a.isoformat(), b.isoformat())
    try:
        orders = client.orders(a, b)
        now = _local_now(core)
        flattened = []
        observations = []
        for order in orders:
            for item in (order.get("orderItems") or []) if isinstance(order, dict) else []:
                if not isinstance(item, dict):
                    continue
                observations.append({
                    "vendor_item_id": item.get("vendorItemId"),
                    "product_name": item.get("productName"),
                    "vendor_item_name": item.get("vendorItemName"),
                    "external_sku_id": item.get("externalSkuId"),
                    "source": "orders",
                })
        with core._conn(db) as con:
            _catalog_upsert(con, observations, now)
            auto_linked = _auto_link_verified_return_aliases(
                con, [x.get("vendor_item_id") for x in observations], now
            )
            mapping = _product_map(con)
            for order in orders:
                order_id = _text(order.get("orderId"))
                paid_at, paid_date = _paid_parts(order.get("paidAt"))
                items = order.get("orderItems") or []
                for idx, item in enumerate(items if isinstance(items, list) else []):
                    if not isinstance(item, dict):
                        continue
                    oid = _oid(item.get("vendorItemId"))
                    if not order_id or not oid:
                        continue
                    flattened.append((
                        order_id, idx, paid_at, paid_date, oid, mapping.get(oid),
                        _text(item.get("productName")),
                        _num(item.get("salesQuantity")),
                        _num(item.get("unitSalesPrice", item.get("salesPrice"))),
                        _text(item.get("currency")), _json(item), now,
                    ))
            con.execute(
                "DELETE FROM coupang_rg_order_items WHERE paid_date>=? AND paid_date<=?",
                (a.isoformat(), b.isoformat()),
            )
            con.executemany(
                """INSERT INTO coupang_rg_order_items
                   (order_id,item_index,paid_at,paid_date,vendor_item_id,product_id,
                    product_name,sales_quantity,unit_sales_price,currency,raw_json,synced_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                flattened,
            )
        matched = sum(1 for row in flattened if row[5] is not None)
        message = (
            f"주문 {len(orders):,}건 · 상품행 {len(flattened):,}개 저장"
            f" · 검증된 반품옵션 자동연결 {auto_linked:,}개"
        )
        _run_finish(core, db, run_id, "success", len(flattened), matched, message)
        return {"run_id": run_id, "orders": len(orders), "rows": len(flattened), "matched": matched}
    except Exception as exc:
        _run_finish(core, db, run_id, "failed", message=str(exc))
        raise


def _inventory_values(row: dict[str, Any]) -> tuple[float, float]:
    details = row.get("inventoryDetails") if isinstance(row.get("inventoryDetails"), dict) else {}
    sales = row.get("salesCountMap") if isinstance(row.get("salesCountMap"), dict) else {}
    return (
        _num(details.get("totalOrderableQuantity")),
        _num(sales.get("SALES_COUNT_LAST_THIRTY_DAYS")),
    )


def _reconcile_warehouse_inventory(
    core: Any,
    con: sqlite3.Connection,
    warehouse_name: str,
    targets: dict[int, dict[str, Any]],
    run_id: int,
    stock_type: str,
):
    if not (_exists(con, "warehouses") and _exists(con, "inventory_txns")):
        return {"adjusted_rows": 0, "adjusted_qty": 0.0}
    warehouse = con.execute(
        "SELECT id FROM warehouses WHERE name=?", (warehouse_name,)
    ).fetchone()
    if not warehouse:
        if targets:
            raise ValueError(f"{warehouse_name} 창고를 찾지 못했습니다.")
        return {"adjusted_rows": 0, "adjusted_qty": 0.0}
    wid = int(warehouse["id"])
    current = {
        int(r["product_id"]): _num(r["qty"])
        for r in con.execute(
            """SELECT product_id,COALESCE(SUM(qty_delta),0) qty
               FROM inventory_txns WHERE warehouse_id=? GROUP BY product_id""",
            (wid,),
        )
    }
    now = _local_now(core)
    today = date.today().isoformat()
    adjusted_rows = 0
    adjusted_qty = 0.0
    for pid, values in sorted(targets.items()):
        target = _num(values.get("qty"))
        option_ids = sorted({_oid(x) for x in values.get("option_ids", []) if _oid(x)})
        live = current.get(int(pid), 0.0)
        delta = float(target) - live
        if abs(delta) <= 1e-9:
            continue
        is_return = stock_type == "return"
        txn_type = "쿠팡API반품재고조정" if is_return else "쿠팡API재고조정"
        stock_label = "반품상품" if is_return else "새상품"
        source_text = ",".join(option_ids)
        con.execute(
            """INSERT INTO inventory_txns
               (txn_date,product_id,warehouse_id,qty_delta,txn_type,ref_no,memo,created_at)
               VALUES(?,?,?,?,?,?,?,?)""",
            (
                today, int(pid), wid, delta, txn_type,
                f"COUPANG-API-INVENTORY-{run_id}-{stock_type}-{pid}",
                f"쿠팡 API {stock_label} 주문 가능 재고 {_integer(target):,}개로 대사"
                + (f" (옵션ID {source_text})" if source_text else ""),
                now,
            ),
        )
        current[int(pid)] = float(target)
        adjusted_rows += 1
        adjusted_qty += delta
    return {"adjusted_rows": adjusted_rows, "adjusted_qty": adjusted_qty}


def sync_inventory(core: Any, client: CoupangClient, db_path=None):
    db = db_path or core.DEFAULT_DB
    run_id = _run_start(core, db, "inventory", None, None)
    try:
        rows = client.inventory()
        now = _local_now(core)
        targets = []
        with core._conn(db) as con:
            observations = [{
                "vendor_item_id": row.get("vendorItemId"),
                "external_sku_id": row.get("externalSkuId"),
                "source": "inventory",
            } for row in rows if isinstance(row, dict)]
            _catalog_upsert(con, observations, now)
            auto_linked = _auto_link_verified_return_aliases(
                con, [x.get("vendor_item_id") for x in observations], now
            )
            official_normal = _normal_option_ids(con)
            normal_mapping = _normal_product_map(con, active_only=True)
            return_mapping = _return_product_map(con)
            for row in rows:
                oid = _oid(row.get("vendorItemId"))
                if not oid:
                    continue
                qty, sales_30d = _inventory_values(row)
                if oid in official_normal:
                    product_id, stock_type = normal_mapping.get(oid), "normal"
                elif oid in return_mapping:
                    product_id, stock_type = return_mapping[oid], "return"
                elif oid in normal_mapping:
                    product_id, stock_type = normal_mapping[oid], "normal"
                else:
                    product_id, stock_type = None, "unmatched"
                targets.append((oid, product_id, stock_type, qty, sales_30d, row))
            # This table is a current snapshot.  Replace only after every API page
            # was received successfully, so partial failures never erase old data.
            con.execute("DELETE FROM coupang_rg_inventory")
            con.executemany(
                """INSERT INTO coupang_rg_inventory
                   (vendor_item_id,product_id,external_sku_id,orderable_qty,sales_30d,
                    raw_json,synced_at,stock_type)
                   VALUES(?,?,?,?,?,?,?,?)""",
                [(
                    oid, pid, _text(raw.get("externalSkuId")), qty, sales30,
                    _json(raw), now, stock_type,
                ) for oid, pid, stock_type, qty, sales30, raw in targets],
            )
            con.executemany(
                """INSERT INTO coupang_rg_inventory_snapshots
                   (run_id,vendor_item_id,product_id,orderable_qty,sales_30d,captured_at,stock_type)
                   VALUES(?,?,?,?,?,?,?)""",
                [(run_id, oid, pid, qty, sales30, now, stock_type)
                 for oid, pid, stock_type, qty, sales30, _raw in targets],
            )
            normal_targets: dict[int, dict[str, Any]] = {}
            return_targets: dict[int, dict[str, Any]] = {}
            for oid, pid, stock_type, qty, _sales30, _raw in targets:
                if pid is None or stock_type == "unmatched":
                    continue
                bucket = return_targets if stock_type == "return" else normal_targets
                value = bucket.setdefault(int(pid), {"qty": 0.0, "option_ids": []})
                value["qty"] += float(qty)
                value["option_ids"].append(oid)
            normal_adjusted = _reconcile_warehouse_inventory(
                core, con, "쿠팡RG", normal_targets, run_id, "normal"
            )
            return_adjusted = _reconcile_warehouse_inventory(
                core, con, "반품창고", return_targets, run_id, "return"
            )
        matched = sum(1 for _oidv, pid, _kind, _q, _s, _r in targets if pid is not None)
        adjusted = {
            "adjusted_rows": normal_adjusted["adjusted_rows"] + return_adjusted["adjusted_rows"],
            "adjusted_qty": normal_adjusted["adjusted_qty"] + return_adjusted["adjusted_qty"],
            "normal_adjusted_rows": normal_adjusted["adjusted_rows"],
            "normal_adjusted_qty": normal_adjusted["adjusted_qty"],
            "return_adjusted_rows": return_adjusted["adjusted_rows"],
            "return_adjusted_qty": return_adjusted["adjusted_qty"],
        }
        message = (
            f"재고 {len(targets):,}개 저장 · 새상품 조정 {adjusted['normal_adjusted_rows']:,}개"
            f" · 반품상품 조정 {adjusted['return_adjusted_rows']:,}개"
            f" · 검증된 반품옵션 자동연결 {auto_linked:,}개"
        )
        _run_finish(core, db, run_id, "success", len(targets), matched, message)
        return {
            "run_id": run_id, "rows": len(targets), "matched": matched,
            "auto_linked": auto_linked,
            **adjusted,
        }
    except Exception as exc:
        _run_finish(core, db, run_id, "failed", message=str(exc))
        raise


def sync_revenue(core: Any, client: CoupangClient, start: date | str, end: date | str, db_path=None):
    db = db_path or core.DEFAULT_DB
    a, b = _to_date(start), _to_date(end)
    run_id = _run_start(core, db, "revenue", a.isoformat(), b.isoformat())
    try:
        transactions = client.revenue(a, b)
        now = _local_now(core)
        flattened = []
        observations = []
        for transaction in transactions:
            for item in (transaction.get("items") or []) if isinstance(transaction, dict) else []:
                if not isinstance(item, dict):
                    continue
                observations.append({
                    "vendor_item_id": item.get("vendorItemId"),
                    "product_name": item.get("productName"),
                    "vendor_item_name": item.get("vendorItemName"),
                    "external_sku_id": item.get("externalSkuId"),
                    "source": "revenue",
                })
        with core._conn(db) as con:
            _catalog_upsert(con, observations, now)
            auto_linked = _auto_link_verified_return_aliases(
                con, [x.get("vendor_item_id") for x in observations], now
            )
            mapping = _product_map(con)
            for transaction_index, transaction in enumerate(transactions):
                order_id = _text(transaction.get("orderId"))
                sale_type = _text(transaction.get("saleType")) or "UNKNOWN"
                recognition_date = _text(transaction.get("recognitionDate"))[:10]
                items = transaction.get("items") or []
                for idx, item in enumerate(items if isinstance(items, list) else []):
                    if not isinstance(item, dict):
                        continue
                    oid = _oid(item.get("vendorItemId"))
                    if not order_id or not recognition_date or not oid:
                        continue
                    flattened.append((
                        order_id, sale_type, recognition_date, transaction_index, idx,
                        _text(transaction.get("saleDate"))[:10],
                        _text(transaction.get("settlementDate"))[:10],
                        oid, mapping.get(oid), _text(item.get("productName")),
                        _text(item.get("vendorItemName")), _num(item.get("quantity")),
                        _num(item.get("salePrice")), _num(item.get("coupangDiscountCoupon")),
                        _num(item.get("sellerDiscountCoupon")), _num(item.get("downloadableCoupon")),
                        _num(item.get("saleAmount")), _num(item.get("serviceFee")),
                        _num(item.get("serviceFeeVat")), _num(item.get("settlementAmount")),
                        _json({"transaction": transaction, "item": item}), now,
                    ))
            con.execute(
                "DELETE FROM coupang_revenue_items WHERE recognition_date>=? AND recognition_date<=?",
                (a.isoformat(), b.isoformat()),
            )
            con.executemany(
                """INSERT INTO coupang_revenue_items
                   (order_id,sale_type,recognition_date,transaction_index,item_index,sale_date,settlement_date,
                    vendor_item_id,product_id,product_name,vendor_item_name,quantity,sale_price,
                    coupang_discount_coupon,seller_discount_coupon,downloadable_coupon,sale_amount,
                    service_fee,service_fee_vat,settlement_amount,raw_json,synced_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                flattened,
            )
        matched = sum(1 for row in flattened if row[8] is not None)
        message = (
            f"매출 거래 {len(transactions):,}건 · 상품행 {len(flattened):,}개 저장"
            f" · 검증된 반품옵션 자동연결 {auto_linked:,}개"
        )
        _run_finish(core, db, run_id, "success", len(flattened), matched, message)
        return {"run_id": run_id, "transactions": len(transactions), "rows": len(flattened), "matched": matched}
    except Exception as exc:
        _run_finish(core, db, run_id, "failed", message=str(exc))
        raise


def sync_settlements(core: Any, client: CoupangClient, month: str, db_path=None):
    db = db_path or core.DEFAULT_DB
    month = str(month)
    try:
        datetime.strptime(month, "%Y-%m")
    except Exception:
        raise ValueError("정산월은 YYYY-MM 형식이어야 합니다.") from None
    run_id = _run_start(core, db, "settlement", month, month)
    try:
        rows = client.settlements(month)
        now = _local_now(core)
        saved = []
        for idx, row in enumerate(rows):
            row_month = _text(row.get("revenueRecognitionYearMonth")) or month
            saved.append((
                row_month, _text(row.get("settlementType")) or "UNKNOWN",
                _text(row.get("settlementDate")), idx,
                _text(row.get("revenueRecognitionDateFrom"))[:10],
                _text(row.get("revenueRecognitionDateTo"))[:10],
                _num(row.get("totalSale")), _num(row.get("serviceFee")),
                _num(row.get("settlementTargetAmount")), _num(row.get("settlementAmount")),
                _num(row.get("lastAmount")), _num(row.get("pendingReleasedAmount")),
                _num(row.get("deductionAmount")), _num(row.get("finalAmount")),
                _text(row.get("status")), _json(row), now,
            ))
        with core._conn(db) as con:
            con.execute("DELETE FROM coupang_settlement_histories WHERE revenue_month=?", (month,))
            con.executemany(
                """INSERT INTO coupang_settlement_histories
                   (revenue_month,settlement_type,settlement_date,item_index,
                    recognition_date_from,recognition_date_to,total_sale,service_fee,
                    settlement_target_amount,settlement_amount,last_amount,pending_released_amount,
                    deduction_amount,final_amount,status,raw_json,synced_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                saved,
            )
        _run_finish(core, db, run_id, "success", len(saved), len(saved), f"지급내역 {len(saved):,}건 저장")
        return {"run_id": run_id, "rows": len(saved), "matched": len(saved)}
    except Exception as exc:
        _run_finish(core, db, run_id, "failed", message=str(exc))
        raise


def _summary(core: Any, db: Any) -> dict[str, Any]:
    ensure_schema(core, db)
    with core._conn(db) as con:
        last = con.execute(
            """SELECT sync_type,status,rows_received,rows_matched,message,completed_at
               FROM coupang_api_sync_runs ORDER BY id DESC LIMIT 20"""
        ).fetchall()
        counts = {}
        for table in (
            "coupang_rg_order_items", "coupang_rg_inventory",
            "coupang_revenue_items", "coupang_settlement_histories",
        ):
            counts[table] = int(con.execute(f'SELECT COUNT(*) n FROM "{table}"').fetchone()["n"])
        unmatched = {
            "orders": int(con.execute(
                "SELECT COUNT(*) n FROM coupang_rg_order_items WHERE product_id IS NULL"
            ).fetchone()["n"]),
            "inventory": int(con.execute(
                "SELECT COUNT(*) n FROM coupang_rg_inventory WHERE product_id IS NULL"
            ).fetchone()["n"]),
            "revenue": int(con.execute(
                "SELECT COUNT(*) n FROM coupang_revenue_items WHERE product_id IS NULL"
            ).fetchone()["n"]),
        }
        inventory = con.execute(
            """SELECT i.vendor_item_id,i.product_id,i.external_sku_id,
                      i.orderable_qty,i.sales_30d,i.synced_at,i.stock_type,
                      COALESCE(NULLIF(c.vendor_item_name,''),NULLIF(c.product_name,''),'') api_product_name,
                      COALESCE(p.name,'') erp_product_name
               FROM coupang_rg_inventory i
               LEFT JOIN coupang_api_option_catalog c ON c.vendor_item_id=i.vendor_item_id
               LEFT JOIN products p ON p.id=i.product_id
               ORDER BY CASE i.stock_type WHEN 'normal' THEN 0 WHEN 'return' THEN 1 ELSE 2 END,
                        i.orderable_qty DESC,i.vendor_item_id"""
        ).fetchall()
        inventory_types = {
            str(r["stock_type"]): {
                "options": int(r["options"] or 0),
                "qty": _num(r["qty"]),
            }
            for r in con.execute(
                """SELECT stock_type,COUNT(*) options,COALESCE(SUM(orderable_qty),0) qty
                   FROM coupang_rg_inventory GROUP BY stock_type"""
            )
        }
        normal_registry_count = int(con.execute(
            "SELECT COUNT(*) n FROM coupang_normal_option_registry"
        ).fetchone()["n"])
        revenue = con.execute(
            """SELECT substr(recognition_date,1,7) month,
                      SUM(CASE WHEN sale_type='REFUND' THEN -ABS(sale_amount) ELSE sale_amount END) sales,
                      SUM(CASE WHEN sale_type='REFUND' THEN -ABS(service_fee+service_fee_vat)
                               ELSE service_fee+service_fee_vat END) fee,
                      SUM(CASE WHEN sale_type='REFUND' THEN -ABS(settlement_amount)
                               ELSE settlement_amount END) settlement
               FROM coupang_revenue_items GROUP BY substr(recognition_date,1,7)
               ORDER BY month DESC LIMIT 12"""
        ).fetchall()
        settlements = con.execute(
            """SELECT revenue_month,settlement_type,settlement_date,total_sale,
                      service_fee,settlement_target_amount,final_amount,status,synced_at
               FROM coupang_settlement_histories
               ORDER BY revenue_month DESC,settlement_date DESC,item_index DESC LIMIT 30"""
        ).fetchall()
    return {
        "last": [dict(x) for x in last], "counts": counts, "unmatched": unmatched,
        "inventory": [dict(x) for x in inventory],
        "unmatched_inventory": [dict(x) for x in inventory if x["product_id"] is None],
        "normal_unmapped_inventory": [
            dict(x) for x in inventory
            if x["product_id"] is None and x["stock_type"] == "normal"
        ],
        "return_mapping_candidates": [
            dict(x) for x in inventory
            if x["product_id"] is None and x["stock_type"] != "normal"
        ],
        "inventory_types": inventory_types,
        "normal_registry_count": normal_registry_count,
        "revenue": [dict(x) for x in revenue],
        "settlements": [dict(x) for x in settlements],
    }


def _return_mapping_products(core: Any, db: Any) -> list[dict[str, Any]]:
    with core._conn(db) as con:
        if not _exists(con, "products"):
            return []
        cols = _columns(con, "products")
        if not {"id", "name", "option_id"}.issubset(cols):
            return []
        code_expr = "item_code" if "item_code" in cols else "'' AS item_code"
        cost_expr = "unit_cost" if "unit_cost" in cols else "0 AS unit_cost"
        active_where = "WHERE COALESCE(active,1)=1" if "active" in cols else ""
        aliases = set(_return_product_map(con))
        rows = con.execute(
            f"SELECT id,name,option_id,{code_expr},{cost_expr} "
            f"FROM products {active_where} ORDER BY name,id"
        ).fetchall()
        out = []
        for r in rows:
            oid = _oid(r["option_id"])
            code = _text(r["item_code"])
            placeholder = (
                code.upper() in {oid.upper(), ("CP-" + oid).upper()}
                and abs(_num(r["unit_cost"])) <= 1e-9
            )
            if not oid or oid in aliases or placeholder:
                continue
            out.append({
                "id": int(r["id"]),
                "name": _text(r["name"]),
                "option_id": oid,
                "item_code": code,
            })
        return out


def _month_bounds(month: str) -> tuple[date, date]:
    year, month_number = (int(x) for x in str(month).split("-"))
    return (
        date(year, month_number, 1),
        date(year, month_number, calendar.monthrange(year, month_number)[1]),
    )


def _revenue_month_coverage(core: Any, db: Any, month: str) -> dict[str, Any]:
    """Coverage is based on successful manual API runs, including empty days."""
    start, end = _month_bounds(month)
    covered: set[date] = set()
    with core._conn(db) as con:
        rows = con.execute(
            """SELECT period_start,period_end
               FROM coupang_api_sync_runs
               WHERE sync_type='revenue' AND status='success'
                 AND period_end>=? AND period_start<=?""",
            (start.isoformat(), end.isoformat()),
        ).fetchall()
        for row in rows:
            try:
                left = max(start, _to_date(row["period_start"]))
                right = min(end, _to_date(row["period_end"]))
            except Exception:
                continue
            cursor = left
            while cursor <= right:
                covered.add(cursor)
                cursor += timedelta(days=1)
        unmatched = int(con.execute(
            """SELECT COUNT(*) n FROM coupang_revenue_items
               WHERE recognition_date>=? AND recognition_date<=? AND product_id IS NULL""",
            (start.isoformat(), end.isoformat()),
        ).fetchone()["n"])
        row_count = int(con.execute(
            """SELECT COUNT(*) n FROM coupang_revenue_items
               WHERE recognition_date>=? AND recognition_date<=?""",
            (start.isoformat(), end.isoformat()),
        ).fetchone()["n"])
    expected = (end - start).days + 1
    return {
        "complete": len(covered) == expected,
        "covered_days": len(covered),
        "expected_days": expected,
        "unmatched": unmatched,
        "rows": row_count,
    }


def _api_revenue_aggregate(core: Any, db: Any, month: str):
    start, end = _month_bounds(month)
    with core._conn(db) as con:
        rows = con.execute(
            """SELECT product_id,
                      SUM(CASE WHEN sale_type='REFUND' THEN -ABS(quantity)
                               ELSE ABS(quantity) END) qty,
                      SUM(CASE WHEN sale_type='REFUND' THEN -ABS(sale_amount)
                               ELSE sale_amount END) realized_sales,
                      SUM(CASE WHEN sale_type='REFUND'
                               THEN -ABS(service_fee + service_fee_vat)
                               ELSE ABS(service_fee + service_fee_vat) END) commission
               FROM coupang_revenue_items
               WHERE recognition_date>=? AND recognition_date<=? AND product_id IS NOT NULL
               GROUP BY product_id""",
            (start.isoformat(), end.isoformat()),
        ).fetchall()
        products = {
            int(r["id"]): {
                "option_id": _oid(r["option_id"]),
                "name": _text(r["name"]),
                "unit_cost": _num(r["unit_cost"]),
            }
            for r in con.execute(
                "SELECT id,option_id,name,unit_cost FROM products"
            )
        }
    return [
        {
            "product_id": int(r["product_id"]),
            "qty": _num(r["qty"]),
            "realized_sales": _num(r["realized_sales"]),
            "commission": max(0.0, _num(r["commission"])),
            **products.get(int(r["product_id"]), {}),
        }
        for r in rows
    ]


def _overlay_confirmed_month(core: Any, db: Any, month: str, original_result: Any):
    """Replace only sales/commission after a complete, fully matched API month.

    Existing Excel-derived RG logistics, return costs and ad settlement remain in
    place.  This prevents API/Excel revenue from being added together twice.
    """
    try:
        import pandas as pd

        base, meta = original_result
        meta = dict(meta or {})
        coverage = _revenue_month_coverage(core, db, month)
        if not coverage["complete"] or coverage["unmatched"] or not coverage["rows"]:
            return base, meta
        api_rows = _api_revenue_aggregate(core, db, month)
        if not api_rows:
            return base, meta

        expected_cols = [
            "product_id", "option_id", "qty", "realized_sales", "cogs",
            "commission", "inout", "delivery", "return_pickup", "return_restock",
        ]
        if base is None or not isinstance(base, pd.DataFrame):
            base = pd.DataFrame()
        if not base.empty and "product_id" not in base.columns:
            # Without a product key, replacing only revenue while retaining exact
            # logistics/return attribution cannot be done safely.
            return base, meta

        # Preserve existing product-level non-sales costs.  Multiple legacy rows
        # for one product are collapsed before the API sales facts are applied.
        preserved: dict[int, dict[str, Any]] = {}
        if not base.empty:
            for _, row in base.iterrows():
                try:
                    pid = int(row.get("product_id"))
                except Exception:
                    continue
                target = preserved.setdefault(pid, {
                    "product_id": pid,
                    "option_id": _oid(row.get("option_id")),
                    "qty": 0.0,
                    "realized_sales": 0.0,
                    "cogs": 0.0,
                    "commission": 0.0,
                    "inout": 0.0,
                    "delivery": 0.0,
                    "return_pickup": 0.0,
                    "return_restock": 0.0,
                })
                for col in ("cogs", "inout", "delivery", "return_pickup", "return_restock"):
                    target[col] += abs(_num(row.get(col)))

        for row in api_rows:
            pid = int(row["product_id"])
            target = preserved.setdefault(pid, {
                "product_id": pid,
                "option_id": row.get("option_id", ""),
                "qty": 0.0,
                "realized_sales": 0.0,
                "cogs": 0.0,
                "commission": 0.0,
                "inout": 0.0,
                "delivery": 0.0,
                "return_pickup": 0.0,
                "return_restock": 0.0,
            })
            target["option_id"] = row.get("option_id") or target.get("option_id", "")
            target["qty"] = _num(row.get("qty"))
            target["realized_sales"] = _num(row.get("realized_sales"))
            target["commission"] = _num(row.get("commission"))
            if target["cogs"] <= 0:
                target["cogs"] = max(0.0, target["qty"]) * _num(row.get("unit_cost"))

        # A legacy row that has no sale in the fully synchronized API month can
        # still carry RG/return costs.  Keep that cost row with zero revenue.
        result = pd.DataFrame(list(preserved.values()), columns=expected_cols)
        if result.empty:
            return base, meta

        revenue = float(pd.to_numeric(result["realized_sales"], errors="coerce").fillna(0).sum())
        cogs = float(pd.to_numeric(result["cogs"], errors="coerce").fillna(0).abs().sum())
        commission = float(pd.to_numeric(result["commission"], errors="coerce").fillna(0).abs().sum())
        rg = sum(
            float(pd.to_numeric(result[col], errors="coerce").fillna(0).abs().sum())
            for col in ("inout", "delivery", "return_pickup", "return_restock")
        )
        ad = abs(_num(meta.get("ad_billable_total")))
        meta["overall_profit"] = revenue - cogs - commission - rg - ad
        meta["api_revenue_source"] = True
        meta["api_revenue_coverage"] = f"{coverage['covered_days']}/{coverage['expected_days']}"
        return result, meta
    except Exception:
        # The legacy confirmed calculation is always the safe fallback.
        return original_result


def _patch_confirmed_pnl(core: Any, db: Any) -> None:
    # Store the unwrapped legacy functions on core.  app.py reloads this module
    # after an in-app update while the Streamlit process stays alive; rebuilding
    # from these originals activates the new module code without wrapper stacking.
    original_confirmed = getattr(core, "_rg_coupang_api_confirmed_original", None)
    if original_confirmed is None:
        original_confirmed = getattr(core, "confirmed_monthly_pnl", None)
        core._rg_coupang_api_confirmed_original = original_confirmed
    original_months = getattr(core, "_rg_coupang_api_monthly_available_original", None)
    if original_months is None:
        original_months = getattr(core, "monthly_available", None)
        core._rg_coupang_api_monthly_available_original = original_months
    if not callable(original_confirmed):
        return

    @functools.wraps(original_confirmed)
    def confirmed_monthly_pnl(month, *args, **kwargs):
        original_result = original_confirmed(month, *args, **kwargs)
        return _overlay_confirmed_month(core, db, str(month), original_result)

    core.confirmed_monthly_pnl = confirmed_monthly_pnl

    if callable(original_months):
        @functools.wraps(original_months)
        def monthly_available(*args, **kwargs):
            values = {str(x) for x in (original_months(*args, **kwargs) or []) if x}
            try:
                with core._conn(db) as con:
                    candidates = [
                        str(r["month"])
                        for r in con.execute(
                            """SELECT DISTINCT substr(recognition_date,1,7) month
                               FROM coupang_revenue_items ORDER BY month DESC"""
                        )
                        if r["month"]
                    ]
                for month in candidates:
                    coverage = _revenue_month_coverage(core, db, month)
                    if coverage["complete"] and coverage["unmatched"] == 0 and coverage["rows"]:
                        values.add(month)
            except Exception:
                pass
            return sorted(values, reverse=True)

        core.monthly_available = monthly_available

    core._rg_coupang_api_confirmed_v09140 = True


def _money(value: Any) -> str:
    return f"{_integer(value):,}원"


def _result_message(label: str, result: dict[str, Any]) -> str:
    rows = int(result.get("rows") or 0)
    matched = int(result.get("matched") or 0)
    extra = ""
    if "normal_adjusted_rows" in result:
        extra = (
            f" · 새상품 재고조정 {int(result.get('normal_adjusted_rows') or 0):,}개"
            f" · 반품상품 재고조정 {int(result.get('return_adjusted_rows') or 0):,}개"
        )
    elif "adjusted_rows" in result:
        extra = f" · ERP 재고조정 {int(result.get('adjusted_rows') or 0):,}개 상품"
    return f"{label} 완료: {rows:,}개 행 저장 · 품목 연결 {matched:,}개{extra}"


def render_page(st: Any, pd: Any, core: Any, db_path=None) -> None:
    """Render the manual-only API page."""
    db = db_path or core.DEFAULT_DB
    ensure_schema(core, db)
    try:
        st.markdown("# 쿠팡 API 연동")
        st.caption("버튼을 누를 때만 쿠팡에서 자료를 가져옵니다. 자동 동기화와 예약 실행은 사용하지 않습니다.")

        saved = None
        load_error = ""
        try:
            saved = load_credentials(core)
        except Exception as exc:
            load_error = str(exc)
        if load_error:
            st.error(load_error)

        with st.expander("API 연결정보", expanded=saved is None):
            st.caption("WING에서 발급한 판매자 ID, Access Key, Secret Key를 입력합니다. Windows 사용자 계정 암호화로 저장됩니다.")
            vendor_id = st.text_input("판매자 ID", value=saved.vendor_id if saved else "", placeholder="A00012345", key="coupang_api_vendor_v09140")
            access_key = st.text_input("Access Key", value=saved.access_key if saved else "", key="coupang_api_access_v09140")
            secret_key = st.text_input("Secret Key", value=saved.secret_key if saved else "", type="password", key="coupang_api_secret_v09140")
            c1, c2 = st.columns(2)
            if c1.button("연결정보 저장", type="primary", use_container_width=True, key="coupang_api_save_v09140"):
                try:
                    save_credentials(core, Credentials(vendor_id, access_key, secret_key))
                    st.success("API 연결정보를 암호화해 저장했습니다.")
                except Exception as exc:
                    st.error(str(exc))
            if c2.button("저장정보 삭제", use_container_width=True, key="coupang_api_delete_v09140"):
                delete_credentials(core)
                st.success("저장된 API 연결정보를 삭제했습니다.")

        with st.expander("정상상품 기준표", expanded=False):
            st.caption(
                "쿠팡 로켓그로스 입고 Excel의 G열 '옵션 ID'와 ERP의 사용 중 상품코드가 "
                "정확히 일치하는 상품만 정상 새상품으로 등록합니다. ERP에 없거나 미사용인 상품은 제외합니다. "
                "이 작업은 쿠팡 API를 호출하지 않습니다."
            )
            inbound_file = st.file_uploader(
                "로켓그로스 입고 Excel",
                type=["xlsx"],
                key="coupang_rg_inbound_options_v09142",
            )
            if st.button(
                "G열 옵션 ID를 정상상품으로 등록",
                type="primary",
                use_container_width=True,
                disabled=inbound_file is None,
                key="coupang_rg_register_normal_v09142",
            ):
                try:
                    parsed = parse_rg_inbound_options(inbound_file)
                    result = register_normal_options(
                        core, parsed, getattr(inbound_file, "name", ""), db
                    )
                    st.success(
                        f"입고표 옵션 {result['rows']:,}개 확인 · "
                        f"ERP 사용 중 상품코드 일치/정상 등록 {result['registered']:,}개 · "
                        f"ERP 미등록·미사용 제외 {result['unmatched']:,}개 · "
                        f"기존 기준 제외 {result['removed']:,}개"
                    )
                    if result["alias_conflicts_removed"]:
                        st.info(
                            "입고표로 정상상품임이 확인되어 기존 반품 연결 "
                            f"{result['alias_conflicts_removed']:,}개를 해제했습니다."
                        )
                except Exception as exc:
                    st.error(str(exc))

        credentials = Credentials(vendor_id, access_key, secret_key) if vendor_id or access_key or secret_key else saved
        ready = credentials is not None
        if ready:
            try:
                credentials.validate()
            except Exception:
                ready = False

        if not ready:
            st.info("연결정보를 입력하고 저장한 뒤 동기화 버튼을 사용하세요.")
            return

        with core._conn(db) as con:
            registry_count = len(_normal_option_ids(con))
        if registry_count == 0:
            st.warning(
                "정상상품 기준표가 비어 있습니다. 재고 동기화 전에 로켓그로스 입고 Excel을 등록하세요. "
                "기준표가 없으면 ERP 상품코드와 직접 일치하는 재고만 새상품으로 처리하고, 반품 자동 판별은 하지 않습니다."
            )

        client = CoupangClient(credentials)
        today = date.today()
        default_start = today - timedelta(days=7)
        st.markdown("### 수동 동기화")
        d1, d2, d3 = st.columns(3)
        start = d1.date_input("조회 시작일", value=default_start, key="coupang_api_start_v09140")
        end = d2.date_input("조회 종료일", value=today, key="coupang_api_end_v09140")
        month = d3.text_input("지급내역 정산월", value=today.strftime("%Y-%m"), key="coupang_api_month_v09140")
        st.caption("긴 기간은 쿠팡 제한에 맞춰 주문 30일, 매출내역 31일 단위로 자동 분할 조회합니다.")

        test_col, blank = st.columns([1, 2])
        if test_col.button("연결 확인", use_container_width=True, key="coupang_api_test_v09140"):
            try:
                # A read-only first request; it does not write any ERP data.
                path = f"/v2/providers/rg_open_api/apis/api/v1/vendors/{credentials.vendor_id}/rg/inventory/summaries"
                payload = client.request(path)
                st.success(f"쿠팡 API 연결 성공 · 첫 응답 {_text(payload.get('message') if isinstance(payload, dict) else 'SUCCESS') or 'SUCCESS'}")
            except Exception as exc:
                st.error(str(exc))

        cols = st.columns(4)
        actions = [
            (cols[0], "주문", "주문 동기화", lambda: sync_orders(core, client, start, end, db)),
            (cols[1], "재고", "재고 동기화", lambda: sync_inventory(core, client, db)),
            (cols[2], "매출·수수료", "매출·수수료 동기화", lambda: sync_revenue(core, client, start, end, db)),
            (cols[3], "지급내역", "지급내역 동기화", lambda: sync_settlements(core, client, month, db)),
        ]
        for col, short, label, action in actions:
            if col.button(label, use_container_width=True, key=f"coupang_api_{short}_v09140"):
                try:
                    with st.spinner(f"{short} 자료를 가져오는 중입니다..."):
                        result = action()
                    st.success(_result_message(short, result))
                except Exception as exc:
                    st.error(str(exc))

        st.caption(
            "처음 연동할 때는 주문·매출의 상품명을 먼저 수집할 수 있도록 "
            "'선택 기간 전체 동기화'를 사용하는 것을 권장합니다."
        )

        if st.button("선택 기간 전체 동기화", type="primary", use_container_width=True, key="coupang_api_all_v09140"):
            completed = []
            try:
                with st.spinner("주문 자료를 가져오는 중입니다..."):
                    completed.append(_result_message("주문", sync_orders(core, client, start, end, db)))
                with st.spinner("매출·수수료 자료를 가져오는 중입니다..."):
                    completed.append(_result_message("매출·수수료", sync_revenue(core, client, start, end, db)))
                with st.spinner("로켓창고 재고를 가져와 ERP 재고와 대사하는 중입니다..."):
                    completed.append(_result_message("재고", sync_inventory(core, client, db)))
                with st.spinner("지급내역을 가져오는 중입니다..."):
                    completed.append(_result_message("지급내역", sync_settlements(core, client, month, db)))
                st.success("전체 동기화를 완료했습니다.\n\n" + "\n\n".join(completed))
            except Exception as exc:
                st.error("완료된 항목: " + (", ".join(x.split(" 완료:")[0] for x in completed) or "없음"))
                st.error(str(exc))

        st.info("광고비·노출·클릭·ROAS는 공개 판매자 API가 없어 기존 광고보고서 Excel 업로드 방식을 계속 사용합니다.")

        summary = _summary(core, db)
        st.markdown("### 연동 현황")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("주문 상품행", f"{summary['counts']['coupang_rg_order_items']:,}개")
        c2.metric("현재 재고 옵션", f"{summary['counts']['coupang_rg_inventory']:,}개")
        c3.metric("매출·수수료 상품행", f"{summary['counts']['coupang_revenue_items']:,}개")
        c4.metric("지급내역", f"{summary['counts']['coupang_settlement_histories']:,}건")
        inv_types = summary.get("inventory_types", {})
        st.caption(
            f"정상상품 기준 옵션 {summary['normal_registry_count']:,}개 · 재고 분류 · "
            f"새상품 {_integer(inv_types.get('normal', {}).get('qty')):,}개 · "
            f"반품상품 {_integer(inv_types.get('return', {}).get('qty')):,}개 · "
            f"미분류 {_integer(inv_types.get('unmatched', {}).get('qty')):,}개"
        )

        missing_total = sum(summary["unmatched"].values())
        if missing_total:
            st.warning(
                "품목관리 옵션ID와 연결되지 않은 API 자료가 있습니다: "
                f"주문 {summary['unmatched']['orders']:,}개 · 재고 {summary['unmatched']['inventory']:,}개 · "
                f"매출 {summary['unmatched']['revenue']:,}개. 정상상품은 품목관리에 등록하고, "
                "자동 판별되지 않은 반품상품은 아래에서 원상품에 연결한 뒤 다시 동기화하세요."
            )

        if summary["normal_unmapped_inventory"]:
            ids = ", ".join(str(x["vendor_item_id"]) for x in summary["normal_unmapped_inventory"][:10])
            suffix = " 외" if len(summary["normal_unmapped_inventory"]) > 10 else ""
            st.warning(
                "입고 Excel로 정상상품임은 확인됐지만 ERP 사용 중 상품코드와 연결되지 않은 옵션이 "
                f"{len(summary['normal_unmapped_inventory']):,}개 있습니다: {ids}{suffix}. "
                "반품으로 처리하지 않으며 품목관리 상품코드와 사용 여부를 확인할 때까지 재고원장은 변경하지 않습니다."
            )

        if summary["return_mapping_candidates"]:
            with st.expander("미분류 재고 옵션을 반품상품으로 연결", expanded=True):
                st.caption(
                    "정상 기준표에 등록된 원상품과 상품명이 하나로 정확히 일치할 때만 자동 판별합니다. "
                    "나머지 중 실제 반품상품 코드만 직접 연결하세요. "
                    "정상상품 자체가 미등록된 경우에는 먼저 품목관리에서 정상 옵션ID를 등록해야 합니다."
                )
                unknown = {str(x["vendor_item_id"]): x for x in summary["return_mapping_candidates"]}
                unknown_ids = list(unknown)
                selected_oid = st.selectbox(
                    "미분류 쿠팡 옵션",
                    unknown_ids,
                    format_func=lambda oid: (
                        f"{oid} · {unknown[oid].get('api_product_name') or '상품명 미확인'}"
                        f" · 재고 {_integer(unknown[oid].get('orderable_qty')):,}개"
                    ),
                    key="coupang_api_unmatched_option_v09141",
                )
                products = _return_mapping_products(core, db)
                if products:
                    by_pid = {int(x["id"]): x for x in products}
                    selected_pid = st.selectbox(
                        "연결할 정상 원상품",
                        list(by_pid),
                        format_func=lambda pid: (
                            f"{by_pid[pid]['name']} · 옵션ID {by_pid[pid]['option_id']}"
                            f" · 품목코드 {by_pid[pid]['item_code']}"
                        ),
                        key="coupang_api_return_parent_v09141",
                    )
                    if st.button(
                        "선택 옵션을 반품상품으로 연결",
                        type="primary",
                        use_container_width=True,
                        key="coupang_api_save_return_mapping_v09141",
                    ):
                        row = unknown[selected_oid]
                        save_return_mapping(
                            core,
                            selected_oid,
                            selected_pid,
                            row.get("api_product_name") or "",
                            db,
                        )
                        st.success(
                            "반품상품 연결을 저장했습니다. 재고 동기화를 다시 누르면 "
                            "해당 옵션 수량이 원상품의 반품창고 재고에 합산됩니다."
                        )
                else:
                    st.info("연결할 정상상품이 없습니다. 품목관리에서 정상상품과 옵션ID를 먼저 등록하세요.")

        if summary["revenue"]:
            st.markdown("#### API 매출·수수료 월별 요약")
            rdf = pd.DataFrame(summary["revenue"])
            rdf = rdf.rename(columns={"month": "월", "sales": "매출인식액", "fee": "판매수수료(VAT포함)", "settlement": "정산대상액"})
            for col in ("매출인식액", "판매수수료(VAT포함)", "정산대상액"):
                rdf[col] = rdf[col].map(_money)
            st.dataframe(rdf, use_container_width=True, hide_index=True)
            try:
                coverage = _revenue_month_coverage(core, db, month)
                if coverage["complete"] and not coverage["unmatched"]:
                    st.success(f"{month} 매출·수수료는 월 전체 {coverage['expected_days']}일 동기화가 완료되어 기존 확정손익에 API 기준으로 반영됩니다.")
                else:
                    st.caption(
                        f"{month} 확정손익 API 반영 조건: 월 전체 동기화 "
                        f"{coverage['covered_days']}/{coverage['expected_days']}일 · "
                        f"미매칭 {coverage['unmatched']:,}개. 조건이 완료되기 전에는 기존 확정자료를 유지합니다."
                    )
            except Exception:
                pass

        if summary["inventory"]:
            with st.expander("최근 로켓창고 재고 보기"):
                idf = pd.DataFrame(summary["inventory"]).rename(columns={
                    "vendor_item_id": "옵션ID", "external_sku_id": "판매자 SKU",
                    "orderable_qty": "주문 가능 재고", "sales_30d": "최근 30일 판매",
                    "synced_at": "동기화 시각", "product_id": "ERP 상품ID",
                    "stock_type": "재고구분", "api_product_name": "쿠팡 상품명",
                    "erp_product_name": "ERP 원상품",
                })
                idf["재고구분"] = idf["재고구분"].map({
                    "normal": "새상품", "return": "반품상품", "unmatched": "미분류",
                }).fillna(idf["재고구분"])
                st.dataframe(idf, use_container_width=True, hide_index=True)

        if summary["settlements"]:
            with st.expander("최근 지급내역 보기"):
                sdf = pd.DataFrame(summary["settlements"]).rename(columns={
                    "revenue_month": "매출인식월", "settlement_type": "정산유형",
                    "settlement_date": "지급일", "total_sale": "총판매액",
                    "service_fee": "판매수수료", "settlement_target_amount": "정산대상액",
                    "final_amount": "최종지급액", "status": "지급상태",
                    "synced_at": "동기화 시각",
                })
                for col in ("총판매액", "판매수수료", "정산대상액", "최종지급액"):
                    sdf[col] = sdf[col].map(_money)
                sdf["정산유형"] = sdf["정산유형"].map({
                    "MONTHLY": "월정산", "WEEKLY": "주정산",
                    "ADDITIONAL": "추가지급", "RESERVE": "최종액지급",
                }).fillna(sdf["정산유형"])
                sdf["지급상태"] = sdf["지급상태"].map({
                    "DONE": "지급완료", "SUBJECT": "지급예정",
                }).fillna(sdf["지급상태"])
                st.dataframe(sdf, use_container_width=True, hide_index=True)

        if summary["last"]:
            with st.expander("최근 동기화 이력", expanded=True):
                labels = {"orders": "주문", "inventory": "재고", "revenue": "매출·수수료", "settlement": "지급내역"}
                ldf = pd.DataFrame(summary["last"]).rename(columns={
                    "sync_type": "구분", "status": "상태", "rows_received": "수신행",
                    "rows_matched": "품목연결", "message": "결과", "completed_at": "완료시각",
                })
                ldf["구분"] = ldf["구분"].map(lambda x: labels.get(str(x), str(x)))
                ldf["상태"] = ldf["상태"].map(lambda x: "완료" if x == "success" else "실패" if x == "failed" else "진행 중")
                st.dataframe(ldf, use_container_width=True, hide_index=True)
    except Exception as exc:
        st.error("쿠팡 API 연동 화면 오류: " + str(exc))


def patch_source(source: str) -> str:
    """Add the API page to the legacy source before grouped-nav AST rewriting."""
    if _MARKER in source:
        return source

    label_line = f'        "{PAGE_LABEL}",\n'
    if label_line not in source:
        anchors = [
            '        "📥  기존ERP 이관",\n',
            '        "📦  재고관리",\n',
        ]
        for anchor in anchors:
            if anchor in source:
                source = source.replace(anchor, anchor + label_line, 1)
                break
        else:
            raise RuntimeError("v0.9.140 쿠팡 API 메뉴를 추가할 위치를 찾지 못했습니다.")

    handler_anchor = (
        "# ------------------------------\n"
        "# Inventory\n"
        "# ------------------------------\n"
        'elif page == "📦  재고관리":\n'
    )
    if handler_anchor not in source:
        raise RuntimeError("v0.9.140 쿠팡 API 화면을 추가할 위치를 찾지 못했습니다.")
    handler = (
        f"{_MARKER}\n"
        "# ------------------------------\n"
        "# Coupang Open API (manual sync only)\n"
        "# ------------------------------\n"
        f'elif page == "{PAGE_LABEL}":\n'
        "    coupang_api_sync_v09140.render_page(st, pd, core)\n\n\n"
    )
    return source.replace(handler_anchor, handler + handler_anchor, 1)


def apply(core: Any, db_path=None) -> None:
    """Create local tables only.  This function never calls Coupang."""
    db = db_path or core.DEFAULT_DB
    ensure_schema(core, db)
    _patch_confirmed_pnl(core, db)
