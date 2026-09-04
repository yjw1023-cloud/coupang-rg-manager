"""RG Manager v0.9.136/v0.9.138/v0.9.159 durable purchase-new-item registration.

Fixes three failure modes in the v0.9.132 new-item flow:
1. A newly confirmed JDS raw item must be committed to the ERP DB immediately.
2. The source-name/detail -> product mapping must survive Streamlit reruns/page moves.
3. v0.9.138 bridges the durable/visible matching choice into the legacy purchase
   selectbox BEFORE it renders, so final confirmation actually writes purchase rows.
4. v0.9.159 preserves purchase Excel option/detail information for new JDS items.
   Current purchase sheets use C=product name and D=option/detail; older F/G layouts
   remain supported as a fallback. When the operator leaves the default new-item
   name unchanged, the detail is appended so size/color variants never collapse
   into identical ERP product names.
"""
from __future__ import annotations

import io
import re
import sys
from typing import Any

from openpyxl import load_workbook

_APPLIED = False
_TABLE = "purchase_source_product_map"


def _norm_text(v: Any) -> str:
    return str(v or "").strip()


def _effective_new_name(source_name: Any, source_detail: Any, requested_name: Any) -> str:
    """Preserve a source option/detail in the ERP name unless the user renamed it."""
    source_name = _norm_text(source_name)
    source_detail = _norm_text(source_detail)
    requested_name = _norm_text(requested_name)
    if not requested_name:
        requested_name = source_name
    if source_detail and requested_name == source_name:
        return f"{source_name} [{source_detail}]"
    return requested_name


def _parse_purchase_excel_option_aware(data: bytes, base_module) -> list[dict[str, Any]]:
    """Read current C/D purchase source columns, with legacy F/G fallback.

    Fixed numeric columns remain W=unit cost and AB=quantity.
    """
    if not data:
        return []
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            for r in range(9, ws.max_row + 1):
                # Current user workbook: C=product name, D=option/detail.
                source_name = _norm_text(ws.cell(r, 3).value)
                source_detail = _norm_text(ws.cell(r, 4).value)

                # Older workbook layout retained for compatibility.
                if not source_name:
                    source_name = _norm_text(ws.cell(r, 6).value)
                    source_detail = _norm_text(ws.cell(r, 7).value)

                unit_cost = ws.cell(r, 23).value
                qty = ws.cell(r, 28).value
                if not source_name:
                    continue
                if unit_cost in (None, "") or qty in (None, ""):
                    continue
                q = base_module._num(qty)
                cost = base_module._num(unit_cost)
                if q == 0:
                    continue
                rows.append(
                    {
                        "index": len(rows) + 1,
                        "sheet": ws.title,
                        "source_row": r,
                        "source_name": source_name,
                        "source_detail": source_detail,
                        "marking": _norm_text(ws.cell(r, 1).value),
                        "qty": q,
                        "unit_cost": cost,
                        "amount": q * cost,
                    }
                )
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return rows


def _ensure_schema(core_module, db_path) -> None:
    core_module.init_db(db_path)
    with core_module._conn(db_path) as con:
        con.execute(
            f"""CREATE TABLE IF NOT EXISTS {_TABLE}(
                   source_name TEXT NOT NULL,
                   source_detail TEXT NOT NULL DEFAULT '',
                   product_id INTEGER NOT NULL,
                   created_at TEXT NOT NULL,
                   PRIMARY KEY(source_name, source_detail)
               )"""
        )
        con.execute(
            f"CREATE INDEX IF NOT EXISTS ix_{_TABLE}_product ON {_TABLE}(product_id)"
        )
        try:
            con.commit()
        except Exception:
            pass


def _next_jds_code_in_connection(con) -> str:
    rows = con.execute("SELECT item_code FROM products WHERE item_code IS NOT NULL").fetchall()
    used_numbers: set[int] = set()
    used_codes: set[str] = set()
    for row in rows:
        code = _norm_text(row["item_code"])
        if not code:
            continue
        used_codes.add(code.upper())
        m = re.fullmatch(r"JDS(\d+)", code, flags=re.IGNORECASE)
        if m:
            n = int(m.group(1))
            if 1 <= n <= 9999:
                used_numbers.add(n)
    for n in range(1, 10000):
        code = f"JDS{n:04d}"
        if n not in used_numbers and code.upper() not in used_codes:
            return code
    raise RuntimeError("JDS0001~JDS9999 품목코드를 모두 사용 중입니다.")


def _mapped_product(con, source_name: str, source_detail: str):
    return con.execute(
        f"""SELECT p.id,p.item_code,p.name
            FROM {_TABLE} m
            JOIN products p ON p.id=m.product_id
            WHERE m.source_name=? AND m.source_detail=?
              AND p.active=1 AND p.option_id IS NULL
            LIMIT 1""",
        (source_name, source_detail),
    ).fetchone()


def _single_exact_raw(con, name: str):
    rows = con.execute(
        """SELECT id,item_code,name
           FROM products
           WHERE active=1 AND option_id IS NULL AND item_type='raw' AND name=?
           ORDER BY id""",
        (name,),
    ).fetchall()
    return rows[0] if len(rows) == 1 else None


def _save_mapping(con, core_module, source_name: str, source_detail: str, product_id: int) -> None:
    con.execute(
        f"""INSERT INTO {_TABLE}(source_name,source_detail,product_id,created_at)
            VALUES(?,?,?,?)
            ON CONFLICT(source_name,source_detail) DO UPDATE SET
              product_id=excluded.product_id,
              created_at=excluded.created_at""",
        (source_name, source_detail, int(product_id), core_module.now_iso()),
    )


def _create_or_reuse_products(patch_module, core_module, db_path, selected_groups):
    """Durably create/reuse selected JDS raw items and persist source mappings."""
    if not selected_groups:
        return {}
    _ensure_schema(core_module, db_path)
    created: dict[tuple[str, str], dict[str, Any]] = {}

    with core_module._conn(db_path) as con:
        for g in selected_groups:
            source_name = _norm_text(g.get("source_name"))
            source_detail = _norm_text(g.get("source_detail"))
            key = (source_name, source_detail)
            name = _effective_new_name(source_name, source_detail, g.get("new_name"))
            if not name:
                raise ValueError("신규 품목명이 비어 있는 항목이 있습니다.")

            row = _mapped_product(con, source_name, source_detail)
            status = "mapped"
            if row is None:
                # A prior attempt may have committed the raw item but lost only the
                # screen/session mapping. Reuse one exact raw-name match so retrying
                # the same purchase file never creates another JDS item.
                row = _single_exact_raw(con, name)
                status = "reused"

            if row is None:
                code = _next_jds_code_in_connection(con)
                cur = con.execute(
                    """INSERT INTO products(item_code,option_id,name,item_type,unit_cost,active,updated_at)
                       VALUES(?,NULL,?,'raw',?,1,?)""",
                    (
                        code,
                        name,
                        float(g.get("avg_unit_cost") or 0),
                        core_module.now_iso(),
                    ),
                )
                pid = int(cur.lastrowid)
                row = {"id": pid, "item_code": code, "name": name}
                status = "created"

            pid = int(row["id"])
            _save_mapping(con, core_module, source_name, source_detail, pid)
            created[key] = {
                "product_id": pid,
                "item_code": _norm_text(row["item_code"]),
                "name": _norm_text(row["name"]),
                "status": status,
            }

        try:
            con.commit()
        except Exception:
            pass

    return created


def _saved_mappings(core_module, db_path, excel_rows) -> dict[tuple[str, str], int]:
    _ensure_schema(core_module, db_path)
    keys = {
        (_norm_text(r.get("source_name")), _norm_text(r.get("source_detail")))
        for r in (excel_rows or [])
    }
    if not keys:
        return {}
    out: dict[tuple[str, str], int] = {}
    with core_module._conn(db_path) as con:
        for key in keys:
            row = _mapped_product(con, key[0], key[1])
            if row is not None:
                out[key] = int(row["id"])
    return out


def apply(patch_module, core_module):
    global _APPLIED
    if patch_module is None:
        return patch_module
    if getattr(patch_module, "_rg_purchase_new_item_persist_v09136_applied", False):
        return patch_module

    base = patch_module.base
    original_review = patch_module._render_review_table

    # v0.9.159: current purchase workbook columns are C=product and D=option/detail.
    # Keep old F/G support so older purchase files still load.
    def parse_purchase_excel(data):
        return _parse_purchase_excel_option_aware(data, base)

    base._parse_purchase_excel = parse_purchase_excel

    def create_new_products(core_obj, db_path, selected_groups):
        return _create_or_reuse_products(patch_module, core_obj, db_path, selected_groups)

    def render_review_table(st_obj, pd_obj, core_obj, db_path, file_fp,
                            excel_rows, meta):
        # Restore confirmed source->JDS mappings before drawing the compact table.
        saved = _saved_mappings(core_obj, db_path, excel_rows)
        if saved:
            overrides_all = st_obj.session_state.setdefault(base._OVERRIDE_KEY, {})
            overrides = overrides_all.setdefault(file_fp, {})
            for row in excel_rows:
                key = (
                    _norm_text(row.get("source_name")),
                    _norm_text(row.get("source_detail")),
                )
                pid = saved.get(key)
                if pid is not None:
                    overrides[str(int(row["index"]))] = int(pid)
            overrides_all[file_fp] = overrides
            st_obj.session_state[base._OVERRIDE_KEY] = overrides_all

        return original_review(
            st_obj, pd_obj, core_obj, db_path, file_fp, excel_rows, meta
        )

    patch_module._create_new_products = create_new_products
    patch_module._render_review_table = render_review_table
    base._render_review_table = render_review_table

    _ensure_schema(core_module, core_module.DEFAULT_DB)

    # v0.9.138: the visible compact review runs after the legacy hidden widgets.
    # Install a replacement v0.9.0 presentation wrapper that reads the durable
    # mapping and preselects those legacy widgets BEFORE they render. This is the
    # missing bridge that lets the existing final-confirmation engine actually
    # insert purchase_lines and own-warehouse inventory for newly registered JDS.
    try:
        purchase_module = sys.modules.get("purchase_v06")
        bridge = __import__("purchase_match_state_bridge_v09138", fromlist=["*"])
        bridge.apply(purchase_module, core_module, base, patch_module)
    except Exception as exc:
        print(f"RG Manager v0.9.138 purchase state bridge failed: {exc}", file=sys.stderr)

    patch_module._rg_purchase_new_item_persist_v09136_applied = True
    _APPLIED = True
    return patch_module
