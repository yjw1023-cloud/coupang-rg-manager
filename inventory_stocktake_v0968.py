"""RG Manager v0.9.68 warehouse inventory Excel export + stocktake adjustment.

Workflow
--------
1. Export one Excel workbook with one sheet per warehouse.
2. User enters only `실사수량` for counted rows.
3. Upload the same workbook.
4. ERP re-reads the live inventory at upload time and previews `실사수량 - 현재ERP재고`.
5. After explicit confirmation, post only the difference as `재고실사조정` inventory_txns.

Safety
------
- Inventory is never overwritten directly; every adjustment stays auditable.
- Product/warehouse identifiers are validated against the current DB.
- Same product+warehouse cannot appear twice in one upload.
- Negative physical counts are rejected.
- Exact duplicate files cannot be applied twice.
- If ERP stock changed after the workbook was exported, the preview warns the user
  and adjustment is still calculated from the live ERP balance, not the stale workbook value.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import hashlib
import math
import re
from typing import Any

import pandas as pd
import streamlit as st

_REQUIRED_INVENTORY_VIEW = {"품목코드", "상품명", "반품창고", "자체창고", "쿠팡RG"}
_PRIMARY_WAREHOUSES = ["자체창고", "쿠팡RG", "반품창고"]
_APPLIED = False


def _num(v: Any) -> float:
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return 0.0
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if not v:
                return 0.0
        x = float(v)
        return 0.0 if math.isnan(x) else x
    except Exception:
        return 0.0


def _display_code(item_code: Any, option_id: Any = None) -> str:
    text = "" if item_code is None else str(item_code).strip()
    if re.fullmatch(r"CP-\d+", text):
        return str(option_id or text[3:])
    return text


def _clean_oid(v: Any) -> str:
    if v is None:
        return ""
    try:
        x = float(v)
        if math.isfinite(x) and abs(x - round(x)) < 1e-9:
            return str(int(round(x)))
    except Exception:
        pass
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def _ensure_schema(core, db):
    core.init_db(db)
    with core._conn(db) as con:
        con.execute(
            """CREATE TABLE IF NOT EXISTS inventory_stocktake_imports(
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   file_hash TEXT NOT NULL UNIQUE,
                   file_name TEXT,
                   stocktake_date TEXT NOT NULL,
                   ref_no TEXT NOT NULL UNIQUE,
                   input_rows INTEGER NOT NULL DEFAULT 0,
                   adjusted_rows INTEGER NOT NULL DEFAULT 0,
                   plus_qty REAL NOT NULL DEFAULT 0,
                   minus_qty REAL NOT NULL DEFAULT 0,
                   memo TEXT,
                   created_at TEXT NOT NULL
               )"""
        )


def _master(core, db):
    core.init_db(db)
    with core._conn(db) as con:
        products = con.execute(
            """SELECT id,item_code,option_id,name,item_type,active
               FROM products ORDER BY name,item_code,id"""
        ).fetchall()
        warehouses = con.execute(
            "SELECT id,name FROM warehouses ORDER BY id"
        ).fetchall()
        balances = con.execute(
            """SELECT product_id,warehouse_id,COALESCE(SUM(qty_delta),0) qty
               FROM inventory_txns GROUP BY product_id,warehouse_id"""
        ).fetchall()

    wh_by_id = {int(r["id"]): str(r["name"] or "") for r in warehouses}
    wh_by_name = {str(r["name"] or ""): int(r["id"]) for r in warehouses}
    bal = {(int(r["product_id"]), int(r["warehouse_id"])): _num(r["qty"]) for r in balances}
    any_bal = {}
    for (pid, _wid), q in bal.items():
        any_bal[pid] = any_bal.get(pid, 0.0) + abs(q)

    prod = {}
    for r in products:
        pid = int(r["id"])
        prod[pid] = {
            "id": pid,
            "item_code": str(r["item_code"] or ""),
            "display_code": _display_code(r["item_code"], r["option_id"]),
            "option_id": _clean_oid(r["option_id"]),
            "name": str(r["name"] or ""),
            "item_type": str(r["item_type"] or ""),
            "active": int(r["active"] or 0),
            "has_any_stock": any_bal.get(pid, 0.0) > 1e-12,
        }
    return prod, wh_by_id, wh_by_name, bal


def _ordered_warehouses(wh_by_name: dict[str, int]) -> list[str]:
    known = [x for x in _PRIMARY_WAREHOUSES if x in wh_by_name]
    extras = [x for x in wh_by_name if x not in known]
    return known + extras


def _sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name or "창고"))[:31] or "창고"
    out = base
    n = 2
    while out in used:
        suffix = f"_{n}"
        out = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(out)
    return out


def _build_workbook(core, db) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule
    except Exception as exc:
        raise RuntimeError(f"Excel 생성 모듈(openpyxl)을 불러오지 못했습니다: {exc}")

    products, _wh_by_id, wh_by_name, balances = _master(core, db)
    warehouses = _ordered_warehouses(wh_by_name)
    if not warehouses:
        raise ValueError("등록된 창고가 없습니다.")

    wb = Workbook()
    ws = wb.active
    ws.title = "사용방법"
    ws["A1"] = "RG Manager 재고 실사"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "1. 각 창고 시트의 '실사수량' 열에 실제 확인한 수량만 입력하세요."
    ws["A4"] = "2. 실사하지 않은 행은 빈칸으로 두면 조정 대상에서 제외됩니다."
    ws["A5"] = "3. ERP상품ID·창고·품목코드·쿠팡 옵션ID·상품명·ERP현재고는 수정하지 않는 것을 권장합니다."
    ws["A6"] = "4. 업로드 시 ERP의 현재 재고를 다시 조회한 뒤 차이만 '재고실사조정' 이력으로 반영합니다."
    ws["A7"] = "5. 같은 파일은 중복 적용할 수 없습니다."
    ws["A9"] = "다운로드 시각"
    ws["B9"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 22

    header_fill = PatternFill("solid", fgColor="CFE3FF")
    header_font = Font(bold=True, color="0D3768")
    thin = Side(style="thin", color="D8E0EA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    plus_fill = PatternFill("solid", fgColor="E2F0D9")
    minus_fill = PatternFill("solid", fgColor="FCE4D6")

    used = {"사용방법"}
    headers = [
        "ERP상품ID", "창고", "품목코드", "쿠팡 옵션ID", "상품명", "상태",
        "ERP현재고", "실사수량", "차이", "비고",
    ]

    for wh_name in warehouses:
        wid = wh_by_name[wh_name]
        sh = wb.create_sheet(_sheet_name(wh_name, used))
        sh.append(headers)
        for cell in sh[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        rows = []
        for pid, p in products.items():
            # Active products are always included so a physically found item can be
            # entered even when ERP stock is currently zero. Archived products stay
            # visible only when they still carry stock somewhere.
            if not p["active"] and not p["has_any_stock"]:
                continue
            qty = balances.get((pid, wid), 0.0)
            rows.append((
                p["name"].lower(), p["display_code"], pid,
                [pid, wh_name, p["display_code"], p["option_id"], p["name"],
                 "사용중" if p["active"] else "보관", qty, None, None, None],
            ))
        rows.sort(key=lambda x: (x[0], x[1], x[2]))
        for _name, _code, _pid, values in rows:
            sh.append(values)

        last = max(2, sh.max_row)
        for r in range(2, last + 1):
            sh.cell(r, 9).value = f'=IF(H{r}="","",H{r}-G{r})'
            sh.cell(r, 8).fill = editable_fill
            for c in range(1, 11):
                sh.cell(r, c).border = border
                sh.cell(r, c).alignment = Alignment(vertical="center")
            for c in (7, 8, 9):
                sh.cell(r, c).number_format = '#,##0.##'

        dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        dv.error = "실사수량은 0 이상의 숫자로 입력하세요."
        dv.errorTitle = "실사수량 입력 오류"
        sh.add_data_validation(dv)
        if last >= 2:
            dv.add(f"H2:H{last}")
            sh.conditional_formatting.add(
                f"I2:I{last}", CellIsRule(operator="greaterThan", formula=["0"], fill=plus_fill)
            )
            sh.conditional_formatting.add(
                f"I2:I{last}", CellIsRule(operator="lessThan", formula=["0"], fill=minus_fill)
            )

        sh.freeze_panes = "A2"
        sh.auto_filter.ref = f"A1:J{last}"
        sh.column_dimensions["A"].hidden = True
        widths = {"B": 14, "C": 18, "D": 18, "E": 52, "F": 10, "G": 14, "H": 14, "I": 14, "J": 28}
        for col, width in widths.items():
            sh.column_dimensions[col].width = width
        sh.row_dimensions[1].height = 24

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except Exception:
        pass
    return isinstance(v, str) and not v.strip()


def _parse_file(raw: bytes):
    try:
        xls = pd.ExcelFile(BytesIO(raw), engine="openpyxl")
    except Exception as exc:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {exc}")

    required = {"ERP상품ID", "창고", "실사수량"}
    parsed = []
    skipped_sheets = []
    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
        except Exception as exc:
            raise ValueError(f"'{sheet}' 시트를 읽지 못했습니다: {exc}")
        if df is None or df.empty:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        if not required.issubset(set(df.columns)):
            skipped_sheets.append(sheet)
            continue
        for row_no, (_, r) in enumerate(df.iterrows(), start=2):
            actual_raw = r.get("실사수량")
            if _is_blank(actual_raw):
                continue
            try:
                actual = float(str(actual_raw).replace(",", "").strip())
            except Exception:
                raise ValueError(f"{sheet}!{row_no}행 실사수량이 숫자가 아닙니다: {actual_raw}")
            if not math.isfinite(actual) or actual < 0:
                raise ValueError(f"{sheet}!{row_no}행 실사수량은 0 이상의 숫자여야 합니다.")

            pid_raw = r.get("ERP상품ID")
            try:
                pid = int(round(float(pid_raw)))
            except Exception:
                raise ValueError(f"{sheet}!{row_no}행 ERP상품ID가 올바르지 않습니다: {pid_raw}")
            wh = str(r.get("창고") or "").strip()
            if not wh:
                raise ValueError(f"{sheet}!{row_no}행 창고가 비어 있습니다.")

            parsed.append({
                "sheet": sheet,
                "row_no": row_no,
                "product_id": pid,
                "warehouse": wh,
                "display_code": str(r.get("품목코드") or "").strip(),
                "option_id": _clean_oid(r.get("쿠팡 옵션ID")),
                "name": str(r.get("상품명") or "").strip(),
                "exported_qty": None if _is_blank(r.get("ERP현재고")) else _num(r.get("ERP현재고")),
                "actual_qty": actual,
                "note": str(r.get("비고") or "").strip(),
            })
    if not parsed:
        raise ValueError("'실사수량'이 입력된 행이 없습니다.")
    return parsed, skipped_sheets


def _validate_preview(core, db, parsed):
    products, _wh_by_id, wh_by_name, balances = _master(core, db)
    seen = set()
    rows = []
    warnings = []

    for x in parsed:
        key = (int(x["product_id"]), str(x["warehouse"]))
        if key in seen:
            raise ValueError(
                f"같은 상품과 창고가 두 번 입력되었습니다: ERP상품ID {key[0]} / {key[1]}"
            )
        seen.add(key)

        p = products.get(int(x["product_id"]))
        if not p:
            raise ValueError(
                f"{x['sheet']}!{x['row_no']}행 ERP상품ID {x['product_id']}를 현재 ERP에서 찾을 수 없습니다."
            )
        if x["warehouse"] not in wh_by_name:
            raise ValueError(
                f"{x['sheet']}!{x['row_no']}행 창고 '{x['warehouse']}'를 현재 ERP에서 찾을 수 없습니다."
            )

        if x["option_id"] and p["option_id"] and x["option_id"] != p["option_id"]:
            raise ValueError(
                f"{x['sheet']}!{x['row_no']}행 옵션ID가 ERP상품ID와 일치하지 않습니다. "
                f"엑셀 {x['option_id']} / ERP {p['option_id']}"
            )
        if x["display_code"] and x["display_code"] != p["display_code"]:
            raise ValueError(
                f"{x['sheet']}!{x['row_no']}행 품목코드가 ERP상품ID와 일치하지 않습니다. "
                f"엑셀 {x['display_code']} / ERP {p['display_code']}"
            )
        if x["name"] and x["name"] != p["name"]:
            warnings.append(
                f"{x['warehouse']} · {p['display_code']}: 상품명이 엑셀과 현재 ERP에서 다릅니다. 현재 ERP 상품명으로 처리합니다."
            )

        wid = wh_by_name[x["warehouse"]]
        live = balances.get((p["id"], wid), 0.0)
        exported = x["exported_qty"]
        stale = exported is not None and abs(exported - live) > 1e-9
        delta = x["actual_qty"] - live
        rows.append({
            "product_id": p["id"],
            "warehouse_id": wid,
            "창고": x["warehouse"],
            "품목코드": p["display_code"],
            "쿠팡 옵션ID": p["option_id"],
            "상품명": p["name"],
            "엑셀ERP현재고": exported,
            "현재ERP재고": live,
            "실사수량": x["actual_qty"],
            "조정수량": delta,
            "다운로드후변동": "있음" if stale else "",
            "비고": x["note"],
        })

    return rows, warnings


def _duplicate_file(core, db, file_hash: str):
    _ensure_schema(core, db)
    with core._conn(db) as con:
        row = con.execute(
            """SELECT file_name,stocktake_date,ref_no,adjusted_rows,created_at
               FROM inventory_stocktake_imports WHERE file_hash=?""",
            (str(file_hash),),
        ).fetchone()
    return dict(row) if row else None


def _commit(core, db, file_hash: str, file_name: str, stocktake_date: str, rows, memo: str):
    _ensure_schema(core, db)
    changed = [r for r in rows if abs(_num(r["조정수량"])) > 1e-12]
    plus_qty = sum(max(0.0, _num(r["조정수량"])) for r in changed)
    minus_qty = sum(max(0.0, -_num(r["조정수량"])) for r in changed)
    ref_no = "STOCKTAKE-" + datetime.now().strftime("%Y%m%d%H%M%S%f")
    now = core.now_iso()

    with core._conn(db) as con:
        dup = con.execute(
            "SELECT ref_no FROM inventory_stocktake_imports WHERE file_hash=?",
            (str(file_hash),),
        ).fetchone()
        if dup:
            raise ValueError(f"이미 적용한 동일 파일입니다. 참조번호: {dup['ref_no']}")

        # Re-read live balances inside the write transaction.  If inventory moved
        # since the preview, recompute the delta from the same physical count.
        live_rows = con.execute(
            """SELECT product_id,warehouse_id,COALESCE(SUM(qty_delta),0) qty
               FROM inventory_txns GROUP BY product_id,warehouse_id"""
        ).fetchall()
        live = {(int(r["product_id"]), int(r["warehouse_id"])): _num(r["qty"]) for r in live_rows}

        adjusted_rows = 0
        plus_qty = 0.0
        minus_qty = 0.0
        for r in rows:
            pid = int(r["product_id"])
            wid = int(r["warehouse_id"])
            current = live.get((pid, wid), 0.0)
            actual = _num(r["실사수량"])
            delta = actual - current
            if abs(delta) <= 1e-12:
                continue
            row_memo = f"재고실사 · {file_name}"
            if memo.strip():
                row_memo += f" · {memo.strip()}"
            if str(r.get("비고") or "").strip():
                row_memo += f" · {str(r['비고']).strip()}"
            con.execute(
                """INSERT INTO inventory_txns
                   (txn_date,product_id,warehouse_id,qty_delta,txn_type,ref_no,memo,created_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (
                    str(stocktake_date), pid, wid, delta, "재고실사조정", ref_no,
                    row_memo, now,
                ),
            )
            adjusted_rows += 1
            plus_qty += max(0.0, delta)
            minus_qty += max(0.0, -delta)

        con.execute(
            """INSERT INTO inventory_stocktake_imports
               (file_hash,file_name,stocktake_date,ref_no,input_rows,adjusted_rows,
                plus_qty,minus_qty,memo,created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                str(file_hash), str(file_name), str(stocktake_date), ref_no,
                len(rows), adjusted_rows, plus_qty, minus_qty, str(memo or ""), now,
            ),
        )

    return {
        "ref_no": ref_no,
        "input_rows": len(rows),
        "adjusted_rows": adjusted_rows,
        "plus_qty": plus_qty,
        "minus_qty": minus_qty,
    }


def _recent(core, db):
    _ensure_schema(core, db)
    with core._conn(db) as con:
        rows = con.execute(
            """SELECT stocktake_date,file_name,ref_no,input_rows,adjusted_rows,
                      plus_qty,minus_qty,created_at
               FROM inventory_stocktake_imports ORDER BY id DESC LIMIT 10"""
        ).fetchall()
    return pd.DataFrame([dict(r) for r in rows])


def _fmt_qty(v: Any) -> str:
    x = _num(v)
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x)):,}"
    return f"{x:,.2f}".rstrip("0").rstrip(".")


def _render_tools(core, db, dataframe_fn):
    st.markdown("### 창고별 재고 엑셀 · 재고실사")
    st.caption(
        "창고별 현재고를 엑셀로 내려받아 실사수량을 입력한 뒤 다시 업로드하면, "
        "ERP 현재고와의 차이만 재고실사조정 이력으로 반영합니다."
    )

    c1, c2 = st.columns([1, 1])
    try:
        export_bytes = _build_workbook(core, db)
        c1.download_button(
            "📥 창고별 재고 실사 엑셀 다운로드",
            data=export_bytes,
            file_name=f"RG_창고별_재고실사_{date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="rg_inventory_stocktake_download_v0968",
        )
    except Exception as exc:
        c1.error(f"재고 엑셀 생성 실패: {exc}")

    c2.info("엑셀에서는 노란색 '실사수량' 열만 입력하세요. 빈칸은 조정하지 않습니다.")

    seq = int(st.session_state.get("_rg_stocktake_upload_seq_v0968", 0))
    with st.expander("📤 재고 실사 엑셀 업로드 및 조정", expanded=False):
        upload = st.file_uploader(
            "실사수량을 입력한 엑셀 파일",
            type=["xlsx"],
            key=f"rg_inventory_stocktake_upload_v0968_{seq}",
        )
        if upload is None:
            recent = _recent(core, db)
            if not recent.empty:
                st.caption("최근 재고실사 적용 이력")
                show = recent.rename(columns={
                    "stocktake_date": "실사일", "file_name": "파일", "ref_no": "참조번호",
                    "input_rows": "입력행", "adjusted_rows": "조정행",
                    "plus_qty": "증가수량", "minus_qty": "감소수량", "created_at": "적용시각",
                })
                dataframe_fn(show, use_container_width=True, hide_index=True)
            return

        raw = upload.getvalue()
        file_hash = hashlib.sha256(raw).hexdigest()
        duplicate = _duplicate_file(core, db, file_hash)
        if duplicate:
            st.warning(
                f"이미 적용한 동일 파일입니다. {duplicate.get('stocktake_date')} · "
                f"{duplicate.get('ref_no')} · 조정 {int(duplicate.get('adjusted_rows') or 0):,}행"
            )
            return

        try:
            parsed, skipped = _parse_file(raw)
            rows, warnings = _validate_preview(core, db, parsed)
        except Exception as exc:
            st.error(str(exc))
            return

        preview = pd.DataFrame(rows)
        changed = preview[pd.to_numeric(preview["조정수량"], errors="coerce").fillna(0).abs() > 1e-12].copy()
        stale_count = int((preview["다운로드후변동"] == "있음").sum())
        plus = float(pd.to_numeric(changed["조정수량"], errors="coerce").clip(lower=0).sum()) if not changed.empty else 0.0
        minus = float((-pd.to_numeric(changed["조정수량"], errors="coerce").clip(upper=0)).sum()) if not changed.empty else 0.0

        st.success(
            f"실사수량 입력 {len(preview):,}행 · 실제 조정 {len(changed):,}행 · "
            f"재고 증가 {_fmt_qty(plus)} · 재고 감소 {_fmt_qty(minus)}"
        )
        if stale_count:
            st.warning(
                f"엑셀을 내려받은 뒤 ERP 재고가 변한 행이 {stale_count:,}개 있습니다. "
                "조정수량은 엑셀의 예전 ERP현재고가 아니라 지금 ERP의 현재고를 기준으로 계산했습니다."
            )
        for msg in warnings[:5]:
            st.warning(msg)
        if skipped:
            st.caption("실사 양식이 아닌 시트는 건너뛰었습니다: " + ", ".join(skipped))

        view_cols = [
            "창고", "품목코드", "쿠팡 옵션ID", "상품명", "엑셀ERP현재고",
            "현재ERP재고", "실사수량", "조정수량", "다운로드후변동", "비고",
        ]
        dataframe_fn(preview[view_cols], use_container_width=True, hide_index=True, height=min(520, 38 * (len(preview) + 1)))

        if changed.empty:
            st.info("현재 ERP 재고와 실사수량이 모두 같아서 조정할 재고가 없습니다.")
            return

        d1, d2 = st.columns([1, 2])
        stocktake_date = d1.date_input(
            "재고실사 기준일", value=date.today(), key=f"rg_stocktake_date_v0968_{seq}"
        )
        memo = d2.text_input(
            "메모(선택)", placeholder="예: 8월 자체창고 정기 실사", key=f"rg_stocktake_memo_v0968_{seq}"
        )
        confirm = st.checkbox(
            f"위 {len(changed):,}개 재고 조정 내용을 확인했습니다.",
            key=f"rg_stocktake_confirm_v0968_{seq}",
        )
        if st.button(
            "재고실사 조정 적용",
            type="primary",
            disabled=not confirm,
            key=f"rg_stocktake_commit_v0968_{seq}",
        ):
            try:
                result = _commit(
                    core, db, file_hash, upload.name, stocktake_date.isoformat(), rows, memo
                )
            except Exception as exc:
                st.error(f"재고실사 조정 실패: {exc}")
                return
            st.session_state["_rg_stocktake_success_v0968"] = (
                f"재고실사 조정 완료 · {result['ref_no']} · "
                f"조정 {result['adjusted_rows']:,}행 · 증가 {_fmt_qty(result['plus_qty'])} · "
                f"감소 {_fmt_qty(result['minus_qty'])}"
            )
            st.session_state["_rg_stocktake_upload_seq_v0968"] = seq + 1
            st.rerun()


def apply(core_module):
    global _APPLIED
    if _APPLIED or getattr(st, "_rg_inventory_stocktake_v0968_applied", False):
        return

    previous_dataframe = st.dataframe

    def dataframe_wrapper(data=None, *args, **kwargs):
        if (
            isinstance(data, pd.DataFrame)
            and _REQUIRED_INVENTORY_VIEW.issubset(set(data.columns))
            and not {"구분", "상태"}.issubset(set(data.columns))
        ):
            db = core_module.DEFAULT_DB
            success = st.session_state.pop("_rg_stocktake_success_v0968", None)
            if success:
                st.success(success)
            try:
                _render_tools(core_module, db, previous_dataframe)
            except Exception as exc:
                st.error(f"재고 엑셀/실사 기능 오류: {exc}")
        return previous_dataframe(data, *args, **kwargs)

    st.dataframe = dataframe_wrapper
    st._rg_inventory_stocktake_v0968_applied = True
    _APPLIED = True
